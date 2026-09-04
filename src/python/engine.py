# elips — Editor de LIcitacions PúbliqueS
# Copyright (C) 2026  Francesc Rambla i Marigot
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import re
import json
import os
from datetime import datetime, date, time, timedelta
from collections import defaultdict
from decimal import Decimal

from openpyxl import load_workbook
from jinja2 import Environment, StrictUndefined, DebugUndefined, Undefined, pass_context
from jinja2.exceptions import UndefinedError, TemplateSyntaxError

def sanitize_empty_jinja_tags(src):
    if not src:
        return src
    src = re.sub(r'\\{\\{\\s*\\}\\}', '[Variable sense nom]', src)
    src = re.sub(r'\\{\\%\\s*\\%\\}', '', src)
    return src

_INLINE_TRAILING_BLOCK_TAG_RE = re.compile(
    r'\{%-?\s*(for|if|elif|else|endfor|endif)\b(?:(?!%\}).)*?(-?%\}|\+%\})'
)
_INLINE_VAR_TAG_RE = re.compile(r'\{\{.*?\}\}')

def protect_inline_trailing_block_tags(template_src):
    """
    trim_blocks=True strips the newline immediately after any {% ... %}
    tag — correct when the tag stands alone on its own line (e.g. a
    DYNAMIC_TABLE's {% for %}/{% endfor %} rows), but wrong when the tag
    sits at the end of a line that carries real content of its own (e.g.
    TRANSPOSED_TABLE's inline per-row `{% for %}...{% endfor %}` loop):
    there, that line's own trailing newline separates it from the next
    row/paragraph and must survive. Jinja2 lets a single tag opt out of
    trim_blocks with a `+` right before its closing `%}`; this scans every
    line and adds it only where content precedes a tag that closes the
    line, so existing templates (written before this distinction existed)
    render correctly without the author having to know this syntax.
    """
    if not template_src:
        return template_src
    lines = template_src.split('\n')
    out_lines = []
    for line in lines:
        matches = list(_INLINE_TRAILING_BLOCK_TAG_RE.finditer(line))
        if not matches:
            out_lines.append(line)
            continue
        last = matches[-1]
        if line[last.end():].strip() != '':
            out_lines.append(line)
            continue
        if last.group(2) in ('-%}', '+%}'):
            out_lines.append(line)
            continue
        before_tag = _INLINE_TRAILING_BLOCK_TAG_RE.sub('', line[:last.start()])
        before_tag = _INLINE_VAR_TAG_RE.sub('', before_tag)
        if before_tag.strip() == '':
            out_lines.append(line)
            continue
        fixed_tag = last.group(0)[:-2] + '+%}'
        out_lines.append(line[:last.start()] + fixed_tag + line[last.end():])
    return '\n'.join(out_lines)

def sanitize_id(s, allow_dots=False):
    s = '' if s is None else str(s)
    s = s.replace(' ', '_')
    if allow_dots and '.' in s:
        parts = [sanitize_id(p, allow_dots=False) for p in s.split('.')]
        return '.'.join(parts)
    import unicodedata
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Za-z0-9_]', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    if not s:
        s = '_'
    if not re.match(r'^[A-Za-z_]', s):
        s = '_' + s
    return s

def _col_letter(col_idx_1):
    result = ""
    n = col_idx_1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

def _col_index_from_letter(letter):
    idx = 0
    for ch in letter.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx

# -------------------- Excel -> JSON --------------------
def _custom_json_default(o):
    if isinstance(o, (datetime, date)):
        return o.isoformat()
    if isinstance(o, time):
        return o.isoformat()
    if isinstance(o, timedelta):
        return str(o)
    if isinstance(o, Decimal):
        if o == o.to_integral_value():
            return int(o)
        return float(o)
    if isinstance(o, bytes):
        return o.decode('utf-8', errors='ignore')
    return str(o)

def _to_jsonable(v, date_format='iso'):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        return v.isoformat() if date_format == 'iso' else v.strftime('%d/%m/%Y')
    if isinstance(v, time):
        return v.isoformat()
    if isinstance(v, timedelta):
        return str(v)
    if isinstance(v, Decimal):
        if v == v.to_integral_value():
            return int(v)
        return float(v)
    if isinstance(v, bytes):
        return v.decode('utf-8', errors='ignore')
    return str(v)

def _resolve_simple_cell_ref(form_str, wb_data, wb_formula, current_ws_name='', visited=None):
    if visited is None:
        visited = set()
    if not isinstance(form_str, str) or not form_str.startswith('='):
        return False, None
    clean = form_str.strip()
    if clean in visited:
        return True, ''
    visited.add(clean)
    
    m = re.match(r"^=[+]?(?:'([^']+)'|([A-Za-z0-9_\.]+))![$]?([A-Za-z]+)[$]?([0-9]+)$", clean)
    if m:
        target_sheet = m.group(1) or m.group(2)
        col_str = m.group(3).upper()
        row_num = int(m.group(4))
        
        resolved_sheet = target_sheet
        if resolved_sheet not in wb_data.sheetnames:
            if f"OUT_{resolved_sheet}" in wb_data.sheetnames:
                resolved_sheet = f"OUT_{resolved_sheet}"
            elif resolved_sheet.startswith("OUT_") and resolved_sheet[4:] in wb_data.sheetnames:
                resolved_sheet = resolved_sheet[4:]
        
        if resolved_sheet in wb_data.sheetnames:
            col_num = 0
            for char in col_str:
                col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
            ws_d = wb_data[resolved_sheet]
            ws_f = wb_formula[resolved_sheet]
            
            val_d = ws_d.cell(row_num, col_num).value
            val_f = ws_f.cell(row_num, col_num).value
            
            if isinstance(val_f, str) and val_f.startswith('='):
                is_sub, sub_val = _resolve_simple_cell_ref(val_f, wb_data, wb_formula, resolved_sheet, visited)
                if is_sub:
                    return True, sub_val
            
            if val_d is not None and str(val_d).strip() != '':
                return True, val_d
            if val_f is not None and not str(val_f).startswith('='):
                return True, val_f
            return True, ''
    return False, None

def _read_rows(ws_d, ws_f=None, wb_data=None, wb_formula=None, ws_name='', date_format='iso'):
    ws_f = ws_f if ws_f is not None else ws_d
    rows = []
    max_c = ws_d.max_column
    for r in range(1, ws_d.max_row + 1):
        row = []
        for c in range(1, max_c + 1):
            val_d = ws_d.cell(r, c).value
            val_f = ws_f.cell(r, c).value if ws_f else val_d
            
            val = _to_jsonable(val_d, date_format)
            if wb_data and wb_formula and isinstance(val_f, str) and val_f.startswith('='):
                is_link, target_val = _resolve_simple_cell_ref(val_f, wb_data, wb_formula, ws_name)
                if is_link:
                    if target_val not in (None, ''):
                        val = _to_jsonable(target_val, date_format)
                    else:
                        val = ''
                else:
                    if val is None:
                        val = ''
            elif val is None:
                val = _to_jsonable(val_f, date_format) if val_f is not None else ''

            row.append(val)

        if all(v in (None, '') for v in row):
            continue
        rows.append(row)
    return rows

def _is_kv_header(first_row):
    if not first_row or len(first_row) < 2:
        return False
    a = str(first_row[0]).strip().lower() if first_row[0] not in (None,'') else ''
    b = str(first_row[1]).strip().lower() if first_row[1] not in (None,'') else ''
    return (a in ('clau','key') and b in ('valor','value'))

def _validate_and_detect_kind(rows, raw_name=""):
    raw_upper = (raw_name or "").upper()
    valid_pfxs = ('OUT_', 'EXPORT_', 'JSON_')
    is_prefixed = any(raw_upper.startswith(pfx) for pfx in valid_pfxs)
    is_dotted = '.' in raw_name

    if not rows:
        if is_prefixed or is_dotted:
            return 'tabular'
        return 'kv'

    if _is_kv_header(rows[0]):
        return 'kv_header'

    first = rows[0]
    headers_non_empty = [v for v in first if v not in (None, '')]

    # Single-level sheets (without dots in name) with fewer than 3 columns default to KV
    is_tabular = is_dotted or (len(headers_non_empty) >= 3)

    if not is_tabular:
        return 'kv'

    header_row_idx = 0
    if len(headers_non_empty) <= 1 and len(rows) > 1:
        r1_non_empty = [v for v in rows[1] if v not in (None, '')]
        if len(r1_non_empty) >= 2:
            header_row_idx = 1

    header_row = rows[header_row_idx]

    seen_raw_headers = {}
    seen_clean_headers = {}
    valid_count = 0

    for col_idx_0, raw_val in enumerate(header_row):
        if raw_val in (None, ''):
            continue
        
        col_letter = _col_letter(col_idx_0 + 1)

        if not isinstance(raw_val, str):
            raise ValueError(f"El full '{raw_name}' té una estructura tabular però la capçalera '{raw_val}' (columna {col_letter}) no és un text.")

        clean_h = sanitize_id(raw_val)
        if not clean_h:
            raise ValueError(f"El full '{raw_name}' té una estructura tabular però la capçalera '{raw_val}' (columna {col_letter}) no és un text vàlid.")

        if raw_val in seen_raw_headers:
            prev_col_letter = _col_letter(seen_raw_headers[raw_val] + 1)
            raise ValueError(f"El full '{raw_name}' té una estructura tabular però les capçaleres de les columnes {prev_col_letter} i {col_letter} ('{raw_val}') són iguals.")
        
        if clean_h in seen_clean_headers:
            prev_col_letter = _col_letter(seen_clean_headers[clean_h] + 1)
            prev_raw = header_row[seen_clean_headers[clean_h]]
            raise ValueError(f"El full '{raw_name}' té una estructura tabular però les capçaleres de les columnes {prev_col_letter} i {col_letter} ('{prev_raw}' i '{raw_val}') són iguals.")

        seen_raw_headers[raw_val] = col_idx_0
        seen_clean_headers[clean_h] = col_idx_0
        valid_count += 1

    if valid_count == 0:
        raise ValueError(f"El full '{raw_name}' té una estructura tabular però la fila {header_row_idx + 1} no conté cap capçalera vàlida.")

    return 'tabular', header_row_idx

def _parse_kv(rows, start_row=0):
    out = {}
    for rr in rows[start_row:]:
        if not rr:
            continue
        k = rr[0]
        if k in (None, ''):
            continue
        v = rr[1] if len(rr) > 1 else None
        out[sanitize_id(k)] = v
    return out

def _is_value_empty(val):
    if val is None:
        return True
    if isinstance(val, str):
        return val.strip() == ''
    if isinstance(val, (list, dict)):
        return len(val) == 0
    return False

def _is_row_empty(item, visited=None, depth=0, max_depth=15):
    if depth > max_depth:
        return False
    if visited is None:
        visited = set()
    if isinstance(item, (dict, list)):
        item_id = id(item)
        if item_id in visited:
            return True
        visited.add(item_id)

    if not isinstance(item, dict):
        return _is_value_empty(item)
    for k, v in item.items():
        if k in ('editor_metadata', '_hierarchy_schema', '_path'):
            continue
        if isinstance(v, list):
            non_empty_children = [child for child in v if not _is_row_empty(child, visited, depth + 1, max_depth)]
            if non_empty_children:
                return False
        elif isinstance(v, dict):
            if not _is_row_empty(v, visited, depth + 1, max_depth):
                return False
        else:
            if not _is_value_empty(v):
                return False
    return True

def _parse_table(rows, header_row_idx=0):
    if not rows or len(rows) <= header_row_idx:
        return []
    header_row = rows[header_row_idx]
    headers = []
    for col_idx, raw_h in enumerate(header_row):
        if raw_h in (None, ''):
            continue
        clean_h = sanitize_id(str(raw_h))
        if clean_h:
            headers.append((col_idx, clean_h))

    if not headers:
        return []

    out = []
    for rr in rows[header_row_idx + 1:]:
        if all(v in (None, '', 0, 0.0, '0', '0.0', '00:00:00', False) for v in rr):
            continue
        obj = {}
        for col_idx, h_name in headers:
            val = rr[col_idx] if col_idx < len(rr) else None
            obj[h_name] = val if val is not None else ''
        out.append(obj)
    return out

def _parse_sheet(ws_d, ws_f, wb_data, wb_formula, date_format='iso'):
    rows = _read_rows(ws_d, ws_f, wb_data, wb_formula, ws_d.title, date_format)
    kind_res = _validate_and_detect_kind(rows, ws_d.title)
    if isinstance(kind_res, tuple):
        kind, header_row_idx = kind_res
    else:
        kind = kind_res
        header_row_idx = 0

    headers = []
    if rows and len(rows) > header_row_idx:
        headers = [sanitize_id(h) for h in rows[header_row_idx] if h not in (None, '')]
        
    if kind == 'kv_header':
        return 'kv', _parse_kv(rows, start_row=1), headers
    if kind == 'kv':
        return 'kv', _parse_kv(rows, start_row=0), headers
    if kind == 'tabular':
        return 'tabular', _parse_table(rows, header_row_idx=header_row_idx), headers
    return kind, [], headers

def _cast_numeric_strings(obj):
    if isinstance(obj, dict):
        return {k: _cast_numeric_strings(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_cast_numeric_strings(item) for item in obj]
    elif isinstance(obj, str):
        s = obj.strip()
        # Try casting to integer (ignoring leading zeros to preserve phone numbers, zip codes, etc.)
        if (s.isdigit() and (len(s) == 1 or not s.startswith('0'))) or (s.startswith('-') and s[1:].isdigit() and (len(s[1:]) == 1 or not s[1:].startswith('0'))):
            try:
                return int(s)
            except ValueError:
                pass
        # Try casting to float
        if '.' in s:
            try:
                return float(s)
            except ValueError:
                pass
    return obj

def _get_nested_containers(root, path_parts):
    containers = [root]
    for p in path_parts:
        next_containers = []
        for c in containers:
            if isinstance(c, dict):
                if p in c:
                    val = c[p]
                    if isinstance(val, list):
                        next_containers.extend(val)
                    elif isinstance(val, dict):
                        next_containers.append(val)
            elif isinstance(c, list):
                for item in c:
                    if isinstance(item, dict) and p in item:
                        val = item[p]
                        if isinstance(val, list):
                            next_containers.extend(val)
                        elif isinstance(val, dict):
                            next_containers.append(val)
        containers = next_containers
    return containers

def _extract_flat_rows_for_sheet(data, raw_sheet_name, has_prefixed_sheets, valid_prefixes):
    stripped = raw_sheet_name
    if has_prefixed_sheets:
        for pfx in valid_prefixes:
            if raw_sheet_name.upper().startswith(pfx):
                stripped = raw_sheet_name[len(pfx):]
                break
    parts = [sanitize_id(p) for p in stripped.split('.')]
    
    if len(parts) == 1:
        val = data.get(parts[0])
        if isinstance(val, list):
            clean_rows = []
            for r in val:
                if isinstance(r, dict):
                    clean_rows.append({k: v for k, v in r.items() if not isinstance(v, (list, dict))})
            return clean_rows
        return val

    parent_parts = parts[:-1]
    sub_key = parts[-1]
    
    parents = _get_nested_containers(data, parent_parts)
    flat_rows = []
    
    for p_item in parents:
        if not isinstance(p_item, dict):
            continue
        p_ref_key = next(iter(p_item.keys())) if p_item else None
        p_ref_val = p_item.get(p_ref_key) if p_ref_key else None
        is_id_key = any(term in str(p_ref_key or '').lower() for term in ['id', 'codi', 'code', 'ref', 'key', 'num'])
        
        alt_sub_key = sub_key + 's' if not sub_key.endswith('s') else sub_key[:-1]
        children = p_item.get(sub_key)
        if children is None:
            children = p_item.get(alt_sub_key, [])

        if isinstance(children, list):
            for child in children:
                if isinstance(child, dict):
                    row = {}
                    child_lower = {str(k).strip().lower(): k for k in child.keys()}
                    if p_ref_key and p_ref_val is not None and is_id_key:
                        p_lower = str(p_ref_key).strip().lower()
                        if p_lower in child_lower:
                            row[child_lower[p_lower]] = p_ref_val
                        else:
                            row[p_ref_key] = p_ref_val
                    for k, v in child.items():
                        if not isinstance(v, (list, dict)):
                            if k not in row:
                                row[k] = v
                    flat_rows.append(row)
    return flat_rows

def _is_dummy_key(val):
    if val is None:
        return True
    s = str(val).strip()
    return s in ('', '0', '0.0', '0.00', 'None', 'null', 'false', 'FALSE')

def excel_to_json(excel_path, date_format='iso', strict=False):
    wb_data = load_workbook(excel_path, data_only=True)
    wb_formula = load_workbook(excel_path, data_only=False)
    parsed = {}
    sheet_logs = []
    import_inspection = {}
    
    valid_prefixes = ('OUT_', 'JSON_', 'EXPORT_')
    has_prefixed_sheets = any(sheet.upper().startswith(valid_prefixes) for sheet in wb_data.sheetnames if not sheet.startswith('_'))
    
    sheet_order = []
    for raw_name in wb_data.sheetnames:
        if raw_name in ('editor_metadata', 'editormetadata', '_hierarchy_schema', '_hierarchy_metadata') or raw_name.startswith('_sheet_info'):
            sheet_logs.append({
                'name': raw_name,
                'processed': False,
                'kind': 'omès',
                'reason': 'Pestanya interna del sistema'
            })
            continue

        raw_upper = raw_name.upper()
        should_process = True
        if has_prefixed_sheets and not any(raw_upper.startswith(pfx) for pfx in valid_prefixes):
            should_process = False

        if not should_process:
            sheet_logs.append({
                'name': raw_name,
                'processed': False,
                'kind': 'omès',
                'reason': 'Sense prefix de processament (OUT_)'
            })
            continue

        try:
            kind, data, headers = _parse_sheet(wb_data[raw_name], wb_formula[raw_name], wb_data, wb_formula, date_format)
            parsed[raw_name] = (kind, data, headers)
            sheet_order.append(raw_name)

            clean_n = raw_name
            if has_prefixed_sheets:
                for pfx in valid_prefixes:
                    if raw_name.upper().startswith(pfx):
                        clean_n = raw_name[len(pfx):]
                        break

            inspection_rows = []
            if kind == 'tabular' and isinstance(data, list):
                for r_idx, row in enumerate(data, 1):
                    prims = [v for k, v in row.items() if not k.startswith('_') and not isinstance(v, (dict, list))]
                    is_empty = (len(prims) == 0) or all(v in (0, 0.0, '', None, False, '0', '0.0') for v in prims)
                    inspection_rows.append({
                        'index': r_idx,
                        'status': 'discarded' if is_empty else 'kept',
                        'reason': 'Fila buida / ceros de fórmula' if is_empty else 'Conté dades vàlides',
                        'data': row
                    })
            elif kind == 'kv' and isinstance(data, dict):
                for k, v in data.items():
                    is_empty = (v is None or (isinstance(v, str) and v.strip() == ''))
                    inspection_rows.append({
                        'index': k,
                        'status': 'discarded' if is_empty else 'kept',
                        'reason': 'Valor buit' if is_empty else 'Clau amb valor',
                        'data': { 'key': k, 'value': v }
                    })

            import_inspection[raw_name] = {
                'raw_name': raw_name,
                'clean_name': clean_n,
                'kind': kind,
                'headers': headers or [],
                'total_rows': len(inspection_rows),
                'kept_count': sum(1 for r in inspection_rows if r['status'] == 'kept'),
                'discarded_count': sum(1 for r in inspection_rows if r['status'] == 'discarded'),
                'rows': inspection_rows
            }

            if kind == 'tabular':
                n_cols = len(headers) if headers else (len(data[0].keys()) if (isinstance(data, list) and data) else 0)
                n_rows = len(data) if isinstance(data, list) else 0
                sheet_logs.append({
                    'name': raw_name,
                    'processed': True,
                    'kind': 'tabular',
                    'cols': n_cols,
                    'rows': n_rows
                })
            elif kind == 'kv':
                n_pairs = len(data) if isinstance(data, dict) else 0
                sheet_logs.append({
                    'name': raw_name,
                    'processed': True,
                    'kind': 'kv',
                    'pairs': n_pairs
                })
            else:
                sheet_logs.append({
                    'name': raw_name,
                    'processed': True,
                    'kind': str(kind),
                    'pairs': len(data) if isinstance(data, dict) else (len(data) if isinstance(data, list) else 0)
                })
        except Exception as err:
            sheet_logs.append({
                'name': raw_name,
                'processed': False,
                'kind': 'error',
                'error': str(err)
            })

    custom_hierarchy_keys = {}
    if "_hierarchy_metadata" in wb_data.sheetnames:
        try:
            ws_meta = wb_data["_hierarchy_metadata"]
            rows = list(ws_meta.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r and len(r) >= 3 and r[0]:
                        custom_hierarchy_keys[str(r[0]).strip()] = {
                            'parent_key': str(r[1] or '').strip(),
                            'child_key': str(r[2] or '').strip()
                        }
        except Exception:
            pass

    root = {}
    hierarchy_schema = {}

    for raw_name in sheet_order:
        kind, data, headers = parsed[raw_name]
        stripped = raw_name
        if has_prefixed_sheets:
            for pfx in valid_prefixes:
                if raw_name.upper().startswith(pfx):
                    stripped = raw_name[len(pfx):]
                    break
        
        parts = [sanitize_id(p) for p in stripped.split('.')]
        path_str = '.'.join(parts)
        
        parent_is_tabular = (len(parts) > 2)
        ref_k = None
        if parent_is_tabular:
            ref_k = headers[0] if (headers and len(parts) > 1) else (next(iter(data[0].keys())) if (isinstance(data, list) and data and len(parts) > 1) else None)

        fields = []
        if kind == 'tabular':
            if headers:
                fields = [h for h in headers if h != ref_k]
            elif isinstance(data, list) and data:
                fields = [k for k in data[0].keys() if k != ref_k]
        elif isinstance(data, dict):
            ref_k = None
            fields = [k for k, v in data.items() if not isinstance(v, (list, dict))]
            
        # Build recursive schema tree
        curr_schema_dict = hierarchy_schema
        for idx, part in enumerate(parts):
            is_last = (idx == len(parts) - 1)
            sub_path_str = '.'.join(parts[:idx+1])
            if part not in curr_schema_dict:
                curr_schema_dict[part] = {
                    'sheet': raw_name if is_last else '',
                    'data_path': sub_path_str,
                    'kind': kind if is_last else 'tabular',
                    'ref_key': ref_k if is_last else None,
                    'fields': fields if is_last else [],
                    'children': {}
                }
            else:
                curr_schema_dict[part]['data_path'] = sub_path_str
                if is_last:
                    curr_schema_dict[part]['sheet'] = raw_name
                    curr_schema_dict[part]['kind'] = kind
                    curr_schema_dict[part]['ref_key'] = ref_k
                    curr_schema_dict[part]['fields'] = fields
            
            curr_schema_dict = curr_schema_dict[part]['children']
        
        if len(parts) == 1:
            if parts[0] not in root:
                root[parts[0]] = data if data is not None else {}
            else:
                if isinstance(data, dict) and isinstance(root[parts[0]], dict):
                    for k, v in data.items():
                        root[parts[0]][k] = v
                else:
                    root[parts[0]] = data if data is not None else {}
        else:
            parent_parts = parts[:-1]
            sub_key = parts[-1]
            curr_sub_path = '.'.join(parts)
            
            parents = _get_nested_containers(root, parent_parts)
            if not parents:
                # Only initialize path if parent keys don't exist at all in root
                curr_n = root
                can_create = True
                for p in parent_parts:
                    if p not in curr_n:
                        curr_n[p] = {}
                    elif isinstance(curr_n[p], list):
                        can_create = False
                        break
                    elif not isinstance(curr_n[p], dict):
                        can_create = False
                        break
                    curr_n = curr_n[p]
                if can_create:
                    parents = _get_nested_containers(root, parent_parts)
            if not parents:
                continue
                
            for parent in parents:
                if isinstance(parent, dict):
                    if data is None:
                        data_to_set = []
                    else:
                        data_to_set = data
                    
                    if isinstance(data_to_set, list):
                        if not parent or not data_to_set:
                            parent[sub_key] = data_to_set if data_to_set is not None else []
                            continue

                        # Determine if parent sheet is a KV (Key-Value) sheet or single root entity
                        parent_path_str = '.'.join(parent_parts)
                        parent_kind = None
                        for r_name, p_tuple in parsed.items():
                            r_stripped = r_name
                            if has_prefixed_sheets:
                                for pfx in valid_prefixes:
                                    if r_name.upper().startswith(pfx):
                                        r_stripped = r_name[len(pfx):]
                                        break
                            if '.'.join([sanitize_id(p) for p in r_stripped.split('.')]) == parent_path_str:
                                parent_kind = p_tuple[0]
                                break

                        # If parent is a KV entity or single root dictionary (not an item in a tabular parent list), ALL rows belong to parent[sub_key]
                        is_parent_root_dict = (len(parent_parts) == 1 and isinstance(root.get(parent_parts[0]), dict))
                        if parent_kind in ('kv', 'kv_header') or is_parent_root_dict:
                            parent[sub_key] = data_to_set
                            continue
                        
                        sample_child = data_to_set[0]
                        if isinstance(sample_child, dict):
                            # Check explicit user-defined foreign keys first
                            c_custom = custom_hierarchy_keys.get(curr_sub_path) if custom_hierarchy_keys else None
                            if c_custom and c_custom.get('parent_key') and c_custom.get('child_key'):
                                pk = c_custom['parent_key']
                                ck = c_custom['child_key']
                                if pk in parent:
                                    p_val = str(parent[pk]).strip()
                                    matched_children = [
                                        c for c in data_to_set
                                        if isinstance(c, dict) and str(c.get(ck, '')).strip() == p_val
                                    ]
                                    parent[sub_key] = matched_children
                                    continue

                            # Case-insensitive key matching between parent and child row keys
                            parent_keys_lower = {str(k).strip().lower(): k for k in parent.keys() if k and not _is_dummy_key(parent[k])}
                            common_pairs = []
                            for ck in sample_child.keys():
                                ck_lower = str(ck).strip().lower()
                                if ck_lower in parent_keys_lower:
                                    pk = parent_keys_lower[ck_lower]
                                    common_pairs.append((pk, ck))

                            id_pairs = [(pk, ck) for pk, ck in common_pairs if any(term in ck.lower() for term in ['id', 'codi', 'code', 'ref', 'key', 'num'])]
                            matching_pairs = id_pairs if id_pairs else common_pairs
                            
                            if matching_pairs:
                                matched_children = [
                                    c for c in data_to_set
                                    if isinstance(c, dict) and all(str(c.get(ck, '')).strip() == str(parent.get(pk, '')).strip() for pk, ck in matching_pairs)
                                ]
                                parent[sub_key] = matched_children
                            else:
                                parent[sub_key] = []
                        else:
                            parent[sub_key] = []
                    else:
                        parent[sub_key] = data_to_set

    sheet_info_list = []
    for raw_name in sheet_order:
        if raw_name in ('editor_metadata', '_hierarchy_metadata'):
            continue
        kind, data, headers = parsed[raw_name]
        stripped = raw_name
        pfx = ''
        if has_prefixed_sheets:
            for p in valid_prefixes:
                if raw_name.upper().startswith(p):
                    pfx = p
                    stripped = raw_name[len(p):]
                    break
        parts = [sanitize_id(p) for p in stripped.split('.')]
        full_path = '.'.join(parts)
        
        c_custom = custom_hierarchy_keys.get(full_path, {})
        sheet_info_list.append({
            'raw_name': raw_name,
            'prefix': pfx,
            'clean_name': parts[-1],
            'parent_path': '.'.join(parts[:-1]),
            'full_path': full_path,
            'kind': kind,
            'headers': headers or [],
            'parent_ref_key': c_custom.get('parent_key', ''),
            'child_ref_key': c_custom.get('child_key', '')
        })

    meta_list = []
    if 'editor_metadata' in wb_data.sheetnames:
        try:
            ws_meta = wb_data['editor_metadata']
            headers = [str(ws_meta.cell(1, c).value or '').strip() for c in range(1, ws_meta.max_column + 1)]
            for r in range(2, ws_meta.max_row + 1):
                row_vals = [ws_meta.cell(r, c).value for c in range(1, len(headers) + 1)]
                if all(v in (None, '') for v in row_vals):
                    continue
                row_dict = {}
                for h, v in zip(headers, row_vals):
                    if h:
                        if v is None:
                            row_dict[h] = ''
                        elif isinstance(v, (int, float, bool, str)):
                            row_dict[h] = v
                        else:
                            row_dict[h] = str(v)
                if row_dict.get('group') or row_dict.get('element'):
                    if 'multiple' in row_dict:
                        row_dict['multiple'] = bool(row_dict['multiple']) if row_dict['multiple'] not in ('', None, 0, '0', False) else False
                    meta_list.append(row_dict)
        except Exception:
            pass

    existing_sheet_info = []
    if '_sheet_info' in wb_data.sheetnames:
        try:
            ws_si = wb_data['_sheet_info']
            headers = [str(ws_si.cell(1, c).value or '').strip() for c in range(1, ws_si.max_column + 1)]
            for r in range(2, ws_si.max_row + 1):
                row_vals = [ws_si.cell(r, c).value for c in range(1, len(headers) + 1)]
                if all(v in (None, '') for v in row_vals):
                    continue
                row_dict = {}
                for h, v in zip(headers, row_vals):
                    if h:
                        if v is None:
                            row_dict[h] = ''
                        elif isinstance(v, (int, float, bool, str)):
                            row_dict[h] = v
                        else:
                            row_dict[h] = str(v)
                if row_dict.get('clean_name') or row_dict.get('raw_name'):
                    if isinstance(row_dict.get('headers'), str):
                        row_dict['headers'] = [h.strip() for h in row_dict['headers'].split(',') if h.strip()]
                    existing_sheet_info.append(row_dict)
        except Exception:
            pass

    root['_sheet_info'] = existing_sheet_info if existing_sheet_info else sheet_info_list
    root['_sheet_logs'] = sheet_logs
    root['_excel_import_inspection'] = import_inspection
    root['editor_metadata'] = meta_list

    return {
        'data': root,
        'hierarchy_schema': hierarchy_schema
    }

def update_excel_from_json(excel_path, json_str, out_excel_path):
    wb = load_workbook(excel_path)
    data = json.loads(json_str)
    orphan_records = []
    
    valid_prefixes = ('OUT_', 'JSON_', 'EXPORT_')
    has_prefixed_sheets = any(sheet.upper().startswith(valid_prefixes) for sheet in wb.sheetnames if not sheet.startswith('_') and sheet != 'editor_metadata')
    
    internal_sheets = ('editor_metadata', 'editormetadata', '_sheet_info', '_hierarchy_schema', '_hierarchy_metadata', 'headers', 'orfes')

    # Remove internal system phantom sheets only (never delete _sheet_info metadata sheet)
    for s_name in list(wb.sheetnames):
        s_lower = s_name.lower()
        if s_lower in ('_sheet_info.headers', 'headers') or s_lower.startswith('_sheet_info.'):
            del wb[s_name]

    unprefixed_sheets = [s for s in wb.sheetnames if not any(s.upper().startswith(pfx) for pfx in valid_prefixes) and s not in internal_sheets]
    prefixed_sheets = [s for s in wb.sheetnames if any(s.upper().startswith(pfx) for pfx in valid_prefixes) and s not in internal_sheets]

    sheets_to_process = prefixed_sheets if prefixed_sheets else unprefixed_sheets
    for sheet_name in sheets_to_process:
        sheet_name_upper = sheet_name.upper()
        sheet_id = sanitize_id(sheet_name)

        if sheet_id in ('editor_metadata', 'editormetadata', 'orfes'):
            continue
            
        stripped_name = sheet_name
        if has_prefixed_sheets:
            for prefix in valid_prefixes:
                if sheet_name_upper.startswith(prefix):
                    stripped_name = sheet_name[len(prefix):]
                    break
            
        ws = wb[sheet_name]
        sheet_data = _extract_flat_rows_for_sheet(data, sheet_name, has_prefixed_sheets, valid_prefixes)
        if sheet_data is None:
            continue
            
        rows = _read_rows(ws, ws, wb, wb, sheet_name, 'iso')
        kind_res = _validate_and_detect_kind(rows, sheet_name)
        if isinstance(kind_res, tuple):
            kind, header_row_idx = kind_res
        else:
            kind = kind_res
        
        if kind in ('kv', 'kv_header'):
            start_row = 1 if kind == 'kv_header' else 0
            if isinstance(sheet_data, dict):
                existing_keys = set()
                for r in range(ws.max_row, start_row, -1):
                    k_cell = ws.cell(r, 1)
                    k_val = k_cell.value
                    k_str = str(k_val or '').strip()
                    if k_str.startswith('=') and is_simple_link_formula(k_str):
                        t_cell = get_referenced_cell(ws, k_cell)
                        if t_cell is not None:
                            k_val = t_cell.value
                    if k_val not in (None, ''):
                        s_key = sanitize_id(k_val)
                        if s_key in sheet_data:
                            existing_keys.add(s_key)
                            val = sheet_data[s_key]
                            if not isinstance(val, (list, dict)):
                                write_cell_value(ws, r, 2, val, orphan_records)
                        else:
                            ws.delete_rows(r)
                            
                for k_val, val in sheet_data.items():
                    if isinstance(val, (list, dict)):
                        continue
                    # Internal bookkeeping keys (a group's own '_group_label'
                    # header, '_hierarchy_schema', '_sheet_info', '_path', ...)
                    # must never be written out as a real spreadsheet row, the
                    # same way the 'tabular' branch above already excludes
                    # underscore-prefixed columns.
                    if str(k_val).startswith('_'):
                        continue
                    s_key = sanitize_id(k_val)
                    if s_key not in existing_keys:
                        next_row = ws.max_row + 1
                        ws.cell(next_row, 1).value = k_val
                        write_cell_value(ws, next_row, 2, val, orphan_records)
                        
        elif kind == 'tabular' and isinstance(sheet_data, list):
            if not sheet_data:
                continue

            excel_headers = []
            for c in range(1, ws.max_column + 1):
                cell_c = ws.cell(1, c)
                val = cell_c.value
                val_str = str(val or '').strip()
                if val_str.startswith('=') and is_simple_link_formula(val_str):
                    t_cell = get_referenced_cell(ws, cell_c)
                    if t_cell is not None:
                        val = t_cell.value
                excel_headers.append(sanitize_id(val) if val not in (None, '') else '')
            
            active_cols = [sanitize_id(k) for k in sheet_data[0].keys() if k and k not in internal_sheets and not str(k).startswith('_')]
            
            if active_cols:
                for c_idx in range(len(excel_headers) - 1, -1, -1):
                    h = excel_headers[c_idx]
                    if h and h not in active_cols:
                        ws.delete_cols(c_idx + 1)
            
            excel_headers = []
            for c in range(1, ws.max_column + 1):
                cell_c = ws.cell(1, c)
                val = cell_c.value
                val_str = str(val or '').strip()
                if val_str.startswith('=') and is_simple_link_formula(val_str):
                    t_cell = get_referenced_cell(ws, cell_c)
                    if t_cell is not None:
                        val = t_cell.value
                excel_headers.append(sanitize_id(val) if val not in (None, '') else '')
            
            for h in active_cols:
                if h not in excel_headers:
                    new_col_idx = len(excel_headers) + 1
                    ws.cell(1, new_col_idx).value = h
                    excel_headers.append(h)
            
            # Only delete excess rows if they do not contain template formula links
            for r in range(ws.max_row, len(sheet_data) + 1, -1):
                has_formula = any(str(ws.cell(r, c).value or '').strip().startswith('=') for c in range(1, ws.max_column + 1))
                if not has_formula:
                    ws.delete_rows(r)
                
            for r_idx, row_obj in enumerate(sheet_data):
                excel_row = r_idx + 2
                for c_idx, h in enumerate(excel_headers):
                    matched_key = None
                    for key in row_obj.keys():
                        if sanitize_id(key) == h:
                            matched_key = key
                            break
                    if matched_key is not None:
                        val = row_obj[matched_key]
                        if not isinstance(val, (list, dict)):
                            write_cell_value(ws, excel_row, c_idx + 1, val, orphan_records)

    # EXPLICITLY write editor_metadata sheet with all 14 config columns regardless of prefixes!
    meta_data = data.get('editor_metadata') or data.get('editorMetadata') or []
    if 'editor_metadata' in wb.sheetnames:
        ws = wb['editor_metadata']
    else:
        ws = wb.create_sheet(title='editor_metadata')

    ws.delete_rows(1, max(ws.max_row, 1))
    headers = ['group', 'element', 'type', 'options', 'sourceType', 'multiple', 'vectorPath', 'displayField', 'valueField', 'width', 'calcFn', 'calcVector', 'calcTargetCol', 'calcFormula', 'gridRow', 'gridOrder', 'gridFill', 'label', 'groupLayout', 'itemTitleFormula']
    for c_idx, h in enumerate(headers):
        ws.cell(1, c_idx + 1).value = h
    
    for r_idx, row_obj in enumerate(meta_data):
        excel_row = r_idx + 2
        for c_idx, h in enumerate(headers):
            val = row_obj.get(h, '')
            if isinstance(val, list):
                val = ', '.join(str(x) for x in val)
            elif isinstance(val, bool):
                val = int(val)
            write_cell_value(ws, excel_row, c_idx + 1, val)
    ws.sheet_state = 'hidden'

    # EXPLICITLY write _sheet_info sheet as hidden metadata sheet if available
    sheet_info_data = data.get('_sheet_info') or []
    if sheet_info_data and isinstance(sheet_info_data, list):
        if '_sheet_info' in wb.sheetnames:
            ws_info = wb['_sheet_info']
            ws_info.delete_rows(1, max(ws_info.max_row, 1))
        else:
            ws_info = wb.create_sheet(title='_sheet_info')

        info_headers = ['raw_name', 'prefix', 'clean_name', 'parent_path', 'full_path', 'kind', 'headers', 'parent_ref_key', 'child_ref_key']
        for c_idx, h in enumerate(info_headers):
            ws_info.cell(1, c_idx + 1).value = h

        for r_idx, row_obj in enumerate(sheet_info_data):
            excel_row = r_idx + 2
            if isinstance(row_obj, dict):
                for c_idx, h in enumerate(info_headers):
                    val = row_obj.get(h, '')
                    if isinstance(val, list):
                        val = ', '.join(str(x) for x in val)
                    write_cell_value(ws_info, excel_row, c_idx + 1, val)
        ws_info.sheet_state = 'hidden'

    # EXPLICITLY write 'orfes' sheet if any complex formula destination cells were encountered!
    if orphan_records:
        if 'orfes' in wb.sheetnames:
            ws_orfes = wb['orfes']
            ws_orfes.delete_rows(1, max(ws_orfes.max_row, 1))
        else:
            ws_orfes = wb.create_sheet(title='orfes')

        orfes_headers = ['Full', 'Coordenada', 'Fórmula Original', 'Valor No Escrit']
        for c_idx, h in enumerate(orfes_headers):
            ws_orfes.cell(1, c_idx + 1).value = h

        for r_idx, o_rec in enumerate(orphan_records):
            r_num = r_idx + 2
            ws_orfes.cell(r_num, 1).value = str(o_rec.get('Full', ''))
            ws_orfes.cell(r_num, 2).value = str(o_rec.get('Coordenada', ''))
            orig_formula = str(o_rec.get('Fórmula Original', ''))
            if orig_formula.startswith('='):
                orig_formula = "'" + orig_formula
            ws_orfes.cell(r_num, 3).value = orig_formula
            val_no_escrit = o_rec.get('Valor No Escrit', '')
            if isinstance(val_no_escrit, (list, dict)):
                val_no_escrit = json.dumps(val_no_escrit, ensure_ascii=False)
            ws_orfes.cell(r_num, 4).value = val_no_escrit

    wb.save(out_excel_path)
    return len(orphan_records)

def update_excel_hierarchy(excel_path, hierarchy_config_json, out_excel_path):
    wb = load_workbook(excel_path)
    config = json.loads(hierarchy_config_json)
    
    renames = config.get("renames", {}) if isinstance(config, dict) else (config if isinstance(config, dict) else {})
    for old_sheet, new_name in renames.items():
        if old_sheet in wb.sheetnames and new_name and old_sheet != new_name:
            wb[old_sheet].title = new_name
            
    custom_keys = config.get("custom_keys", {}) if isinstance(config, dict) else {}
    if custom_keys:
        if "_hierarchy_metadata" in wb.sheetnames:
            ws_meta = wb["_hierarchy_metadata"]
            ws_meta.delete_rows(1, ws_meta.max_row)
        else:
            ws_meta = wb.create_sheet("_hierarchy_metadata")
        ws_meta.append(["sub_path", "parent_key", "child_key"])
        for sub_path, k_dict in custom_keys.items():
            if isinstance(k_dict, dict) and (k_dict.get("parent_key") or k_dict.get("child_key")):
                ws_meta.append([sub_path, k_dict.get("parent_key", ""), k_dict.get("child_key", "")])
        ws_meta.sheet_state = "hidden"

    wb.save(out_excel_path)
    return True

# -------------------- Jinja: recuperació d'errors --------------------
RE_DOTTED = re.compile(r"\\b([A-Za-z_][A-Za-z0-9_]*(?:\\.[A-Za-z_][A-Za-z0-9_]*)+)\\b")

class TrackedValue:
    def __init__(self, val, path, enable_links=True):
        self.val = val
        self._path = path
        self.enable_links = enable_links

    def __str__(self):
        if self.val is None:
            return ''
        s = str(self.val)
        if self.enable_links and self._path:
            return f'<a href="#dades.{self._path}" class="data-link" data-path="{self._path}">{s}</a>'
        return s

    def __repr__(self):
        return str(self)

    def __html__(self):
        return str(self)

    def __getattr__(self, name):
        if name.startswith('_') or name in ('val', 'enable_links', 'get', 'keys', 'items', 'values'):
            raise AttributeError(name)
        if name in ('value', 'val'):
            return TrackedValue(self.val, f"{self._path}.{name}" if self._path else name, self.enable_links)
        
        # On-the-fly Foreign Key lookup: search global doc for matching row
        if hasattr(self, '_doc_ref') and isinstance(self._doc_ref, dict) and self.val not in (None, ''):
            val_str = str(self.val)
            for sheet_name, sheet_data in self._doc_ref.items():
                if isinstance(sheet_data, list):
                    for row in sheet_data:
                        if isinstance(row, dict):
                            if any(str(v) == val_str for k_v, v in row.items() if not k_v.startswith('_')):
                                if name in row:
                                    return _wrap_tracked(row[name], f"{self._path}.{name}" if self._path else name, self.enable_links, doc_ref=self._doc_ref)
        
        return Placeholder(f"{self._path}.{name}" if self._path else name)

    def __getitem__(self, item):
        try:
            return self.__getattr__(str(item))
        except AttributeError:
            return Placeholder(f"{self._path}.{item}" if self._path else str(item))

    def __bool__(self):
        return bool(self.val)

    def __len__(self):
        if hasattr(self.val, '__len__'):
            return len(self.val)
        return 0

    def __int__(self):
        return int(self.val)

    def __float__(self):
        return float(self.val)

    def __hash__(self):
        return hash(self.val)

    def __eq__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val == o

    def __ne__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val != o

    def __lt__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val < o

    def __le__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val <= o

    def __gt__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val > o

    def __ge__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        return self.val >= o

    def __add__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = self.val + o
        return TrackedValue(res, self._path, self.enable_links)

    def __radd__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = o + self.val
        return TrackedValue(res, self._path, self.enable_links)

    def __sub__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = self.val - o
        return TrackedValue(res, self._path, self.enable_links)

    def __rsub__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = o - self.val
        return TrackedValue(res, self._path, self.enable_links)

    def __mul__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = self.val * o
        return TrackedValue(res, self._path, self.enable_links)

    def __rmul__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = o * self.val
        return TrackedValue(res, self._path, self.enable_links)

    def __truediv__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = self.val / o
        return TrackedValue(res, self._path, self.enable_links)

    def __rtruediv__(self, other):
        o = other.val if isinstance(other, TrackedValue) else other
        res = o / self.val
        return TrackedValue(res, self._path, self.enable_links)

class TrackedDict(dict):
    def __init__(self, d, path, enable_links=True, visited=None):
        super().__init__()
        self._path = path
        self.enable_links = enable_links
        if visited is None:
            visited = set()
        for k, v in d.items():
            sub_path = f"{path}.{k}" if path else k
            self[k] = _wrap_tracked(v, sub_path, enable_links, visited)

    def __getitem__(self, key):
        if key not in self:
            p = f"{self._path}.{key}" if self._path else str(key)
            return Placeholder(p)
        return super().__getitem__(key)

    def __getattr__(self, name):
        if name.startswith('_') or name in ('get', 'keys', 'items', 'values'):
            raise AttributeError(name)
        try:
            return self[name]
        except KeyError:
            raise AttributeError(name)

    def get(self, key, default=None):
        if key not in self:
            p = f"{self._path}.{key}" if self._path else str(key)
            return Placeholder(p)
        return super().get(key, default)

    def __str__(self):
        if '_default_val' in self:
            # Use the raw scalar here, not the wrapped TrackedValue stored in
            # self['_default_val'] — that one already carries its own
            # `<a>`-link wrapper (added when the FK row's field values were
            # wrapped), and re-wrapping its rendered string in another `<a>`
            # would produce invalid nested anchors in the HTML preview.
            raw = super().__getitem__('_default_val')
            raw_val = raw.val if isinstance(raw, TrackedValue) else raw
            val_str = str(raw_val)
            if getattr(self, 'enable_links', True) and getattr(self, '_path', ''):
                clean_path = self._path.lstrip('.')
                return f'<a href="#dades.{clean_path}" class="data-link" title="Anar a la dada: {clean_path}">{val_str}</a>'
            return val_str
        return super().__str__()

class TrackedList(list):
    def __init__(self, lst, path, enable_links=True, visited=None):
        super().__init__()
        self._path = path
        self.enable_links = enable_links
        if visited is None:
            visited = set()
        for idx, item in enumerate(lst):
            sub_path = f"{path}.{idx}"
            self.append(_wrap_tracked(item, sub_path, enable_links, visited))

class Placeholder:
    def __init__(self, path):
        self._path = path

    def __getattr__(self, name):
        if name.startswith('_') or name in ('get', 'keys', 'items', 'values'):
            raise AttributeError(name)
        return Placeholder(f"{self._path}.{name}")

    def __getitem__(self, key):
        return Placeholder(f"{self._path}[{key}]")

    def __iter__(self):
        return iter([Placeholder(f"{self._path}[0]")])

    def __len__(self):
        return 1

    def __str__(self):
        return f"[{self._path}]"

    def __repr__(self):
        return f"[{self._path}]"

    def __html__(self):
        return f"[{self._path}]"

    def __add__(self, other):
        return Placeholder(f"{self._path} + {other}")
    def __radd__(self, other):
        return Placeholder(f"{other} + {self._path}")
    def __sub__(self, other):
        return Placeholder(f"{self._path} - {other}")
    def __rsub__(self, other):
        return Placeholder(f"{other} - {self._path}")
    def __mul__(self, other):
        return Placeholder(f"{self._path} * {other}")
    def __rmul__(self, other):
        return Placeholder(f"{other} * {self._path}")
    def __truediv__(self, other):
        return Placeholder(f"{self._path} / {other}")
    def __rtruediv__(self, other):
        return Placeholder(f"{other} / {self._path}")

def _placeholder(path):
    return Placeholder(path)

def _get_line(src, lineno):
    if not lineno:
        return ''
    lines = src.splitlines()
    if 1 <= lineno <= len(lines):
        return lines[lineno-1]
    return ''

def _parse_missing(e, line_text):
    msg = str(e)
    m = re.search(r"has no attribute '([^']+)'", msg)
    attr = m.group(1) if m else None

    candidates = RE_DOTTED.findall(line_text or '')
    if attr and candidates:
        for c in candidates:
            if c.endswith('.' + sanitize_id(attr)) or c.endswith('.' + attr):
                return c
    if len(candidates) == 1:
        return candidates[0]
    if attr:
        return attr
    return msg

def _ensure_path(ctx, path):
    parts = path.split('.')
    if not parts:
        return None, None
    cur = ctx
    for p in parts[:-1]:
        if p not in cur or not isinstance(cur.get(p), dict):
            cur[p] = {}
        cur = cur[p]
    return cur, parts[-1]

def render_with_recovery(env, template_src, ctx, pass_label, max_fixes=50):
    issues = []
    current_src = template_src
    last_keypath = None
    repeat_count = 0
    
    for _ in range(max_fixes):
        try:
            out = env.from_string(current_src).render(**ctx)
            return out, issues
        except UndefinedError as e:
            lineno = getattr(e, 'lineno', None)
            line_text = _get_line(current_src, lineno)
            keypath = _parse_missing(e, line_text)
            keypath = '.'.join(sanitize_id(p) for p in str(keypath).split('.'))
            
            if keypath == last_keypath:
                repeat_count += 1
                if repeat_count > 3:
                    break
            else:
                last_keypath = keypath
                repeat_count = 0

            issues.append({
                'pass': pass_label,
                'line': lineno,
                'key': keypath,
                'message': f"Clau no definida '{keypath}' resolta amb marcador a la línia {lineno} de la {pass_label} passada."
            })
            
            if '.' in keypath:
                parent, last = _ensure_path(ctx, keypath)
                if parent is not None and last is not None:
                    parent[last] = _placeholder(keypath)
            else:
                ctx[keypath] = _placeholder(keypath)
            continue
        except TemplateSyntaxError as e:
            lineno = getattr(e, 'lineno', None)
            msg = str(e)
            lines = current_src.splitlines()
            issues.append({
                'pass': pass_label,
                'line': lineno,
                'key': 'syntax_error',
                'message': f"Sintaxi Jinja2 no vàlida a la línia {lineno}: '{msg}'. S'ha corregit automàticament."
            })
            if lineno and 1 <= lineno <= len(lines):
                bad_line = lines[lineno - 1]
                lines[lineno - 1] = bad_line.replace('{{', '&#123;&#123;').replace('}}', '&#125;&#125;').replace('{%', '&#123;&#37;').replace('%}', '&#37;&#125;')
                current_src = '\\n'.join(lines)
                continue
            break
        except Exception as e:
            msg = str(e)
            issues.append({
                'pass': pass_label,
                'line': 0,
                'key': 'render_error',
                'message': f"Error de renderitzat ({pass_label}): {msg}"
            })
            break

    # Fallback to non-strict environment if recovery loop gets stuck.
    # trim_blocks/lstrip_blocks=True (matching env_clean/env_html below):
    # without them, a {% for %}/{% endif %}/etc. tag's own newline survives
    # in the output, leaving a blank line before/after every loop iteration
    # and every if/endif — breaking Markdown tables (a blank line ends a
    # table) and adding unwanted spacing between guarded paragraphs.
    env_lax = Environment(undefined=DebugUndefined, autoescape=False, trim_blocks=True, lstrip_blocks=True)
    env_lax.filters.update(env.filters)
    env_lax.globals.update(env.globals)
    try:
        out = env_lax.from_string(current_src).render(**ctx)
        return out, issues
    except Exception:
        return current_src, issues

SIMPLE_LINK_REGEX = re.compile(r"^=[+]?(?:(?:'([^']+)'|([A-Za-z0-9_\.]+))!)?[$]?[A-Za-z]+[$]?[0-9]+$", re.IGNORECASE)
REF_REGEX = re.compile(r"=[+]?(?:'([^']+)'|([A-Za-z0-9_\.]+))!([A-Za-z0-9$]+)", re.IGNORECASE)

def is_simple_link_formula(val_str):
    if not isinstance(val_str, str):
        return False
    s = val_str.strip()
    if not s.startswith('='):
        return False
    return bool(SIMPLE_LINK_REGEX.match(s))

def get_referenced_cell(ws, cell):
    val = str(cell.value or '').strip()
    if not val.startswith('='):
        return None
    m = REF_REGEX.search(val)
    if m:
        target_sheet_name = m.group(1) or m.group(2)
        cell_coord = m.group(3).replace('$', '')
        wb = ws.parent
        
        matched_sheet = None
        target_upper = target_sheet_name.upper()
        
        # 1. Exact or case-insensitive match
        for s in wb.sheetnames:
            if s.upper() == target_upper:
                matched_sheet = s
                break
                
        # 2. Try with or without valid prefixes (OUT_, JSON_, EXPORT_)
        if not matched_sheet:
            prefixes = ('OUT_', 'JSON_', 'EXPORT_')
            clean_target = target_upper
            for pfx in prefixes:
                if clean_target.startswith(pfx):
                    clean_target = clean_target[len(pfx):]
                    break
            
            for s in wb.sheetnames:
                s_clean = s.upper()
                for pfx in prefixes:
                    if s_clean.startswith(pfx):
                        s_clean = s_clean[len(pfx):]
                        break
                if s_clean == clean_target:
                    matched_sheet = s
                    break
                    
        # 3. Create target sheet on-the-fly if missing in workbook
        if not matched_sheet:
            try:
                matched_sheet = target_sheet_name
                wb.create_sheet(title=matched_sheet)
            except Exception:
                pass

        if matched_sheet and matched_sheet in wb.sheetnames:
            from openpyxl.cell.cell import MergedCell
            target_ws = wb[matched_sheet]
            ref_c = target_ws[cell_coord]
            if isinstance(ref_c, MergedCell) or type(ref_c).__name__ == 'MergedCell':
                for rng in target_ws.merged_cells.ranges:
                    if ref_c.coordinate in rng:
                        ref_c = target_ws.cell(rng.min_row, rng.min_col)
                        break
            return ref_c
            
    return None

def write_cell_value(ws, row_idx, col_idx, value, orphan_records=None):
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row_idx, col_idx)
    if isinstance(cell, MergedCell) or type(cell).__name__ == 'MergedCell':
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                cell = ws.cell(rng.min_row, rng.min_col)
                break

    target_cell = cell
    visited = set()

    current_val_str = str(cell.value or '').strip()

    # 1. Auto-propagate reference formula if current cell lacks formula but preceding row in column has a simple link formula
    if (cell.value is None or current_val_str == '' or not current_val_str.startswith('=')) and row_idx > 2:
        for ref_r in range(2, row_idx):
            prev_cell = ws.cell(ref_r, col_idx)
            if is_simple_link_formula(str(prev_cell.value or '')):
                ref_target = get_referenced_cell(ws, prev_cell)
                if ref_target is not None:
                    ref_sheet_name = ref_target.parent.title
                    col_letter = get_column_letter(col_idx)
                    propagated_formula = f"='{ref_sheet_name}'!{col_letter}{row_idx}"
                    cell.value = propagated_formula
                    target_cell = cell
                    current_val_str = propagated_formula
                    break

    # 2. Check if destination cell itself is a Complex Formula (starts with '=' but is NOT a simple 1-to-1 link)
    if current_val_str.startswith('='):
        if not is_simple_link_formula(current_val_str):
            if orphan_records is not None:
                orphan_records.append({
                    'Full': ws.title,
                    'Coordenada': cell.coordinate,
                    'Fórmula Original': current_val_str,
                    'Valor No Escrit': value
                })
            return

    # 3. Traverse simple link reference chain
    while True:
        target_val_str = str(target_cell.value or '').strip()
        if target_val_str.startswith('='):
            if not is_simple_link_formula(target_val_str):
                if orphan_records is not None:
                    orphan_records.append({
                        'Full': target_cell.parent.title,
                        'Coordenada': target_cell.coordinate,
                        'Fórmula Original': target_val_str,
                        'Valor No Escrit': value
                    })
                return

        ref_cell = get_referenced_cell(target_cell.parent, target_cell)
        if ref_cell is None:
            break
        ref_key = f"{ref_cell.parent.title}!{ref_cell.coordinate}"
        if ref_key in visited:
            break
        visited.add(ref_key)
        target_cell = ref_cell

    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False, default=_custom_json_default)

    # Cast value to correct type before writing
    if isinstance(value, str):
        s = value.strip()
        if (s.isdigit() and (len(s) == 1 or not s.startswith('0'))) or (s.startswith('-') and s[1:].isdigit() and (len(s[1:]) == 1 or not s[1:].startswith('0'))):
            try:
                value = int(s)
            except ValueError:
                pass
        elif '.' in s:
            try:
                value = float(s)
            except ValueError:
                pass
                
    if isinstance(target_cell, MergedCell) or type(target_cell).__name__ == 'MergedCell':
        for rng in target_cell.parent.merged_cells.ranges:
            if target_cell.coordinate in rng:
                target_cell = target_cell.parent.cell(rng.min_row, rng.min_col)
                break

    if not (isinstance(target_cell, MergedCell) or type(target_cell).__name__ == 'MergedCell'):
        target_cell.value = value

def _find_sheet_for_group(wb, sheet_name):
    """Resolves a schema group name (e.g. 'nomconjunt' or 'pres.parts') to the
    actual sheet in `wb`, tolerating the OUT_/JSON_/EXPORT_ prefixes the rest
    of the engine uses. The OUT_-prefixed sheet (if any) wins over a bare
    same-named sheet, since callers always pass the app's schema group name —
    which is exactly the OUT_ sheet's name with the prefix stripped — and a
    same-named *source* sheet existing alongside it (the mirror scenario this
    function primarily exists for) must never shadow it."""
    for pfx in ('OUT_', 'JSON_', 'EXPORT_'):
        candidate = f"{pfx}{sheet_name}"
        if candidate in wb.sheetnames:
            return candidate
    if sheet_name in wb.sheetnames:
        return sheet_name
    clean_target = sanitize_id(sheet_name, allow_dots=True)
    for s in wb.sheetnames:
        stripped = s
        for pfx in ('OUT_', 'JSON_', 'EXPORT_'):
            if s.upper().startswith(pfx):
                stripped = s[len(pfx):]
                break
        if sanitize_id(stripped, allow_dots=True) == clean_target:
            return s
    return None

def analyze_mirror_pattern(excel_path, sheet_name):
    """
    Scans a tabular sheet's EXISTING columns to determine whether the sheet
    is a simple, cell-by-cell mirror of another sheet — every non-empty cell
    in every existing column is a plain `=SourceSheet!Cell` link formula (see
    `is_simple_link_formula`), all columns point at the SAME single source
    sheet, each OUT_ column maps to exactly one source column, and every row
    maps to the source with the same constant row offset ("one column per
    source column", so a new column could plausibly be appended the same way).

    Used by the "add a new field/column to a group" flow (GroupConfigModal.vue
    via useGroupMetadata.js's saveGroupConfig) to decide whether to offer
    replicating a newly-added column into the source sheet too, instead of
    silently leaving the new OUT_ column unlinked to whatever sheet actually
    holds the real data (see bug report: OUT_nomconjunt mirroring `nomconjunt`).

    Returns a JSON string:
      {"is_mirror": true, "source_sheet": "...", "row_offset": N,
       "next_source_col": "E", "header_row": 1, "first_data_row": 2,
       "last_data_row": N}
    or
      {"is_mirror": false, "reason": "..."}
    reason is one of: sheet_not_found, no_data_rows, no_formulas_found,
    non_formula_or_complex_cell, unparseable_formula, unparseable_cell_ref,
    multiple_or_no_source_sheets, column_maps_to_multiple_source_columns,
    inconsistent_row_offset, duplicate_source_column_mapping.
    """
    try:
        wb = load_workbook(excel_path)
        target_name = _find_sheet_for_group(wb, sheet_name)
        if not target_name:
            return json.dumps({"is_mirror": False, "reason": "sheet_not_found"})

        ws = wb[target_name]
        if ws.max_row < 2 or ws.max_column < 1:
            return json.dumps({"is_mirror": False, "reason": "no_data_rows"})

        # First pass: only check whether the sheet has ANY formula at all. A
        # sheet with none is simply not attempting to mirror anything — the
        # ordinary case for every brand-new/plain-data sheet — and must be
        # told apart from one that mixes formulas with plain values (a
        # genuinely ambiguous, only-partially-mirrored sheet).
        any_formula = False
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(1, col_idx).value in (None, ''):
                continue
            for row_idx in range(2, ws.max_row + 1):
                val_str = str(ws.cell(row_idx, col_idx).value or '').strip()
                if val_str and is_simple_link_formula(val_str):
                    any_formula = True
                    break
            if any_formula:
                break
        if not any_formula:
            return json.dumps({"is_mirror": False, "reason": "no_formulas_found"})

        source_sheets = set()
        col_source_cols = {}
        row_offsets = set()

        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(1, col_idx).value
            if header_val in (None, ''):
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                val_str = str(cell.value if cell.value is not None else '').strip()
                if val_str == '':
                    continue
                if not is_simple_link_formula(val_str):
                    return json.dumps({
                        "is_mirror": False,
                        "reason": "non_formula_or_complex_cell",
                        "cell": f"{_col_letter(col_idx)}{row_idx}"
                    })
                m = REF_REGEX.search(val_str)
                if not m:
                    return json.dumps({"is_mirror": False, "reason": "unparseable_formula"})
                source_sheets.add(m.group(1) or m.group(2))
                cell_coord = m.group(3).replace('$', '')
                cm = re.match(r'^([A-Za-z]+)([0-9]+)$', cell_coord)
                if not cm:
                    return json.dumps({"is_mirror": False, "reason": "unparseable_cell_ref"})
                col_source_cols.setdefault(col_idx, set()).add(cm.group(1).upper())
                row_offsets.add(int(cm.group(2)) - row_idx)
        if len(source_sheets) != 1:
            return json.dumps({
                "is_mirror": False,
                "reason": "multiple_or_no_source_sheets",
                "sheets": sorted(source_sheets)
            })
        if any(len(cols) != 1 for cols in col_source_cols.values()):
            return json.dumps({"is_mirror": False, "reason": "column_maps_to_multiple_source_columns"})
        if len(row_offsets) != 1:
            return json.dumps({"is_mirror": False, "reason": "inconsistent_row_offset"})

        used_col_nums = [_col_index_from_letter(next(iter(cols))) for cols in col_source_cols.values()]
        if len(used_col_nums) != len(set(used_col_nums)):
            return json.dumps({"is_mirror": False, "reason": "duplicate_source_column_mapping"})

        return json.dumps({
            "is_mirror": True,
            "source_sheet": next(iter(source_sheets)),
            "row_offset": next(iter(row_offsets)),
            "next_source_col": _col_letter(max(used_col_nums) + 1),
            "header_row": 1,
            "first_data_row": 2,
            "last_data_row": ws.max_row,
        })
    except Exception as ex:
        return json.dumps({"is_mirror": False, "reason": f"error: {ex}"})

def apply_mirror_column(excel_path, sheet_name, new_field_name, out_excel_path):
    """
    Given a sheet already confirmed (by the caller, after `analyze_mirror_pattern`
    returned is_mirror=true AND the user explicitly approved it) to be a simple
    mirror of another sheet, appends `new_field_name` as a new column in BOTH
    the source sheet (a plain header cell — its actual data is left for the
    user to fill in directly, same as any other new column in that sheet) and
    the OUT_ sheet (a `=SourceSheet!Col<row>` formula per existing data row,
    replicating the exact convention `analyze_mirror_pattern` detected).
    Re-runs the analysis itself rather than trusting pre-computed coordinates
    from the caller, so this is safe to call on its own.
    Returns a JSON string: {"applied": true, ...} or {"applied": false, "reason": ...}.
    """
    try:
        analysis = json.loads(analyze_mirror_pattern(excel_path, sheet_name))
        if not analysis.get("is_mirror"):
            return json.dumps({"applied": False, "reason": analysis.get("reason", "not_a_mirror")})

        wb = load_workbook(excel_path)
        target_name = _find_sheet_for_group(wb, sheet_name)
        # analysis["source_sheet"] came straight out of the formula text
        # (REF_REGEX), so it is already the exact sheet name — no OUT_-prefix
        # guessing needed (and none wanted: that heuristic exists to resolve
        # schema group names, and would wrongly prefer an OUT_ sheet here).
        source_name = analysis["source_sheet"]
        if source_name not in wb.sheetnames:
            return json.dumps({"applied": False, "reason": "source_sheet_missing"})

        ws = wb[target_name]
        ws_source = wb[source_name]

        new_col_idx = ws.max_column + 1
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(1, c).value or '').strip() == str(new_field_name).strip():
                return json.dumps({"applied": False, "reason": "column_already_exists"})

        source_col_letter = analysis["next_source_col"]
        source_col_idx = _col_index_from_letter(source_col_letter)
        existing_source_header = ws_source.cell(1, source_col_idx).value
        if existing_source_header not in (None, ''):
            return json.dumps({"applied": False, "reason": "source_column_already_occupied"})

        ws.cell(1, new_col_idx).value = new_field_name
        ws_source.cell(1, source_col_idx).value = new_field_name

        row_offset = analysis["row_offset"]
        for row_idx in range(analysis["first_data_row"], analysis["last_data_row"] + 1):
            ws.cell(row_idx, new_col_idx).value = f"='{source_name}'!{source_col_letter}{row_idx + row_offset}"

        wb.save(out_excel_path)
        return json.dumps({
            "applied": True,
            "source_sheet": source_name,
            "source_col": source_col_letter,
            "out_col": _col_letter(new_col_idx),
        })
    except Exception as ex:
        return json.dumps({"applied": False, "reason": f"error: {ex}"})

def create_default_workbook_from_json(json_str, out_path):
    import json
    from openpyxl import Workbook
    try:
        data = json.loads(json_str)
    except Exception:
        data = {}
        
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    internal_keys = ('editor_metadata', 'editormetadata', '_sheet_info', '_hierarchy_schema', '_hierarchy_metadata', 'headers')

    # Create primary data sheets only
    for sheet_name, sheet_content in data.items():
        if sheet_name in internal_keys or str(sheet_name).startswith('_'):
            continue
            
        ws = wb.create_sheet(title=str(sheet_name))
        if isinstance(sheet_content, dict):
            ws.append(["Clau", "Valor"])
            for k, v in sheet_content.items():
                if isinstance(v, (dict, list)) or k in internal_keys or str(k).startswith('_'):
                    continue
                ws.append([str(k), "" if v is None else str(v)])
        elif isinstance(sheet_content, list) and len(sheet_content) > 0:
            headers = [k for k in sheet_content[0].keys() if not isinstance(sheet_content[0][k], (dict, list)) and k not in internal_keys and not str(k).startswith('_')]
            ws.append(headers)
            for row in sheet_content:
                if isinstance(row, dict):
                    ws.append(["" if row.get(h) is None else str(row.get(h)) for h in headers])

    # Write editor_metadata sheet cleanly with all 16 columns
    editor_meta = data.get('editor_metadata') or data.get('editorMetadata') or []
    if editor_meta:
        ws = wb.create_sheet(title='editor_metadata')
        headers = ['group', 'element', 'type', 'options', 'sourceType', 'multiple', 'vectorPath', 'displayField', 'valueField', 'width', 'calcFn', 'calcVector', 'calcTargetCol', 'calcFormula', 'gridRow', 'gridOrder', 'gridFill', 'label', 'groupLayout', 'itemTitleFormula']
        ws.append(headers)
        for row_obj in editor_meta:
            if isinstance(row_obj, dict):
                ws.append(["" if row_obj.get(h) is None else (", ".join(str(x) for x in row_obj[h]) if isinstance(row_obj[h], list) else str(row_obj[h])) for h in headers])
        ws.sheet_state = 'hidden'

    if not wb.sheetnames:
        wb.create_sheet(title="Dades")
        
    wb.save(out_path)

# -------------------- Localized Formatting Filters & Spelling --------------------
def get_locale_from_context(context):
    for key in ('config', 'CONFIG'):
        cfg = context.get(key)
        if isinstance(cfg, dict):
            locale_val = cfg.get('locale')
            if locale_val:
                return str(locale_val).strip()
    return 'ca_ES'

def catalan_number_to_words(number):
    if number == 0:
        return "zero"
    units = ["", "un", "dos", "tres", "quatre", "cinc", "sis", "set", "vuit", "nou"]
    units_f = ["", "una", "dues", "tres", "quatre", "cinc", "sis", "set", "vuit", "nou"]
    teens = ["deu", "onze", "dotze", "tretze", "catorze", "quinze", "setze", "disset", "divuit", "dinou"]
    tens = ["", "", "vint", "trenta", "quaranta", "cinquanta", "seixanta", "setanta", "vuitanta", "noranta"]
    
    def _convert_below_1000(n, feminine=False):
        if n == 0:
            return ""
        res = []
        h = n // 100
        rem = n % 100
        if h > 0:
            if h == 1:
                res.append("cent")
            else:
                res.append(f"{units_f[h] if feminine else units[h]}-cents")
        if rem > 0:
            if rem < 10:
                res.append(units_f[rem] if feminine else units[rem])
            elif rem < 20:
                res.append(teens[rem - 10])
            else:
                t = rem // 10
                u = rem % 10
                if t == 2:
                    if u == 0:
                        res.append("vint")
                    else:
                        res.append(f"vint-i-{units_f[u] if feminine else units[u]}")
                else:
                    if u == 0:
                        res.append(tens[t])
                    else:
                        res.append(f"{tens[t]}-{units_f[u] if feminine else units[u]}")
        return " ".join(res)

    millions = number // 1000000
    rem = number % 1000000
    thousands = rem // 1000
    rem_units = rem % 1000
    
    parts = []
    if millions > 0:
        if millions == 1:
            parts.append("un milió")
        else:
            parts.append(f"{_convert_below_1000(millions)} milions")
    if thousands > 0:
        if thousands == 1:
            parts.append("mil")
        else:
            parts.append(f"{_convert_below_1000(thousands)} mil")
    if rem_units > 0 or not parts:
        parts.append(_convert_below_1000(rem_units))
    return " ".join(parts).strip()

def catalan_currency_to_words(val):
    euros = int(val)
    cents = int(round((val - euros) * 100))
    if euros == 1:
        euros_str = "un euro"
    else:
        euros_str = f"{catalan_number_to_words(euros)} euros"
    if cents == 0:
        return euros_str
    if cents == 1:
        cents_str = "un cèntim"
    else:
        cents_str = f"{catalan_number_to_words(cents)} cèntims"
    return f"{euros_str} amb {cents_str}"

def catalan_date_to_words(d):
    months = ["", "gener", "febrer", "març", "abril", "maig", "juny", "juliol", "agost", "setembre", "octubre", "novembre", "desembre"]
    day = d.day
    month = d.month
    year = d.year
    day_str = "primer" if day == 1 else catalan_number_to_words(day)
    return f"{day_str} de {months[month]} de {catalan_number_to_words(year)}"

def spanish_number_to_words(number):
    if number == 0:
        return "cero"
    units = ["", "uno", "dos", "tres", "cuatro", "cinco", "seis", "siete", "ocho", "nueve"]
    teens = ["diez", "once", "doce", "trece", "catorce", "quince", "dieciseis", "diecisiete", "dieciocho", "diecinueve"]
    tens = ["", "", "veinte", "treinta", "cuarenta", "cincuenta", "sesenta", "setenta", "ochenta", "noventa"]
    hundreds = ["", "cien", "doscientos", "trescientos", "cuatrocientos", "quinientos", "seiscientos", "setecientos", "ochocientos", "novecientos"]
    
    def _convert_below_1000(n):
        if n == 0:
            return ""
        res = []
        h = n // 100
        rem = n % 100
        if h > 0:
            if h == 1 and rem == 0:
                res.append("cien")
            elif h == 1:
                res.append("ciento")
            else:
                res.append(hundreds[h])
        if rem > 0:
            if rem < 10:
                res.append(units[rem])
            elif rem < 20:
                res.append(teens[rem - 10])
            elif rem < 30:
                if rem == 20:
                    res.append("veinte")
                else:
                    res.append(f"veinti{units[rem%10]}")
            else:
                t = rem // 10
                u = rem % 10
                if u == 0:
                    res.append(tens[t])
                else:
                    res.append(f"{tens[t]} y {units[u]}")
        return " ".join(res)

    millions = number // 1000000
    rem = number % 1000000
    thousands = rem // 1000
    rem_units = rem % 1000
    
    parts = []
    if millions > 0:
        if millions == 1:
            parts.append("un millón")
        else:
            parts.append(f"{_convert_below_1000(millions)} millones")
    if thousands > 0:
        if thousands == 1:
            parts.append("mil")
        else:
            t_str = _convert_below_1000(thousands)
            if t_str.endswith("uno"):
                t_str = t_str[:-3] + "ún"
            parts.append(f"{t_str} mil")
    if rem_units > 0 or not parts:
        parts.append(_convert_below_1000(rem_units))
    return " ".join(parts).strip()

def spanish_currency_to_words(val):
    euros = int(val)
    cents = int(round((val - euros) * 100))
    if euros == 1:
        euros_str = "un euro"
    else:
        euros_str = f"{spanish_number_to_words(euros)} euros"
    if cents == 0:
        return euros_str
    if cents == 1:
        cents_str = "un céntimo"
    else:
        cents_str = f"{spanish_number_to_words(cents)} céntimos"
    return f"{euros_str} con {cents_str}"

def spanish_date_to_words(d):
    months = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    day = d.day
    month = d.month
    year = d.year
    return f"{spanish_number_to_words(day)} de {months[month]} de {spanish_number_to_words(year)}"

def english_number_to_words(number):
    if number == 0:
        return "zero"
    units = ["", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine"]
    teens = ["ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", "seventeen", "eighteen", "nineteen"]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]
    
    def _convert_below_1000(n):
        if n == 0:
            return ""
        res = []
        h = n // 100
        rem = n % 100
        if h > 0:
            res.append(f"{units[h]} hundred")
        if rem > 0:
            if rem < 10:
                res.append(units[rem])
            elif rem < 20:
                res.append(teens[rem - 10])
            else:
                t = rem // 10
                u = rem % 10
                if u == 0:
                    res.append(tens[t])
                else:
                    res.append(f"{tens[t]}-{units[u]}")
        return " ".join(res)
        
    millions = number // 1000000
    rem = number % 1000000
    thousands = rem // 1000
    rem_units = rem % 1000
    
    parts = []
    if millions > 0:
        parts.append(f"{_convert_below_1000(millions)} million")
    if thousands > 0:
        parts.append(f"{_convert_below_1000(thousands)} thousand")
    if rem_units > 0 or not parts:
        parts.append(_convert_below_1000(rem_units))
    return " ".join(parts).strip()

def english_currency_to_words(val):
    dollars = int(val)
    cents = int(round((val - dollars) * 100))
    dollars_str = "one dollar" if dollars == 1 else f"{english_number_to_words(dollars)} dollars"
    if cents == 0:
        return dollars_str
    cents_str = "one cent" if cents == 1 else f"{english_number_to_words(cents)} cents"
    return f"{dollars_str} and {cents_str}"

def english_ordinal_words(n):
    words = {
        1: "first", 2: "second", 3: "third", 4: "fourth", 5: "fifth",
        6: "sixth", 7: "seventh", 8: "eighth", 9: "ninth", 10: "tenth",
        11: "eleventh", 12: "twelfth", 13: "thirteenth", 14: "fourteenth",
        15: "fifteenth", 16: "sixteen", 17: "seventeenth", 18: "eighteenth",
        19: "nineteenth", 20: "twentieth", 30: "thirtieth"
    }
    if n in words:
        return words[n]
    if n < 30:
        return f"twenty-{words[n%10]}"
    if n < 40:
        return f"thirty-{words[n%10]}"
    return english_number_to_words(n)

def english_date_to_words(d):
    months = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    return f"{english_ordinal_words(d.day)} of {months[d.month]}, {english_number_to_words(d.year)}"

def _wrap_res(res, orig_path, enable_links):
    if orig_path:
        return TrackedValue(res, orig_path, enable_links)
    return res

@pass_context
def filter_coin(context, value, currency_symbol=None):
    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)
    val_raw = value.val if isinstance(value, TrackedValue) else value
    locale_code = get_locale_from_context(context)
    if val_raw in (None, ''):
        return _wrap_res('', orig_path, enable_links)
    try:
        val = float(val_raw)
    except (ValueError, TypeError):
        return _wrap_res(val_raw, orig_path, enable_links)
    dec_sep = ','
    thousands_sep = '.'
    curr_before = False
    locale_lower = locale_code.lower()
    if any(x in locale_lower for x in ('en_us', 'en_gb', 'en_ca', 'en_au')):
        dec_sep = '.'
        thousands_sep = ','
        curr_before = True
    if currency_symbol is None:
        if 'en_us' in locale_lower:
            currency_symbol = '$'
        elif 'en_gb' in locale_lower:
            currency_symbol = '£'
        else:
            currency_symbol = '€'
    val_str = f"{val:.2f}"
    parts = val_str.split('.')
    integer_part = parts[0]
    decimal_part = parts[1]
    is_neg = integer_part.startswith('-')
    if is_neg:
        integer_part = integer_part[1:]
    thousands_list = []
    while len(integer_part) > 3:
        thousands_list.insert(0, integer_part[-3:])
        integer_part = integer_part[:-3]
    thousands_list.insert(0, integer_part)
    formatted_int = (('-' if is_neg else '') + thousands_sep.join(thousands_list))
    formatted_num = f"{formatted_int}{dec_sep}{decimal_part}"
    if curr_before:
        res = f"{currency_symbol}{formatted_num}"
    else:
        res = f"{formatted_num}{currency_symbol}"
    return _wrap_res(res, orig_path, enable_links)

@pass_context
def filter_number(context, value, precision=2):
    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)
    val_raw = value.val if isinstance(value, TrackedValue) else value
    locale_code = get_locale_from_context(context)
    if val_raw in (None, ''):
        return _wrap_res('', orig_path, enable_links)
    try:
        val = float(val_raw)
        precision = int(precision)
    except (ValueError, TypeError):
        return _wrap_res(val_raw, orig_path, enable_links)
    dec_sep = ','
    thousands_sep = '.'
    locale_lower = locale_code.lower()
    if any(x in locale_lower for x in ('en_us', 'en_gb', 'en_ca', 'en_au')):
        dec_sep = '.'
        thousands_sep = ','
    val_str = f"{val:.{precision}f}"
    parts = val_str.split('.')
    integer_part = parts[0]
    decimal_part = parts[1] if len(parts) > 1 else ''
    is_neg = integer_part.startswith('-')
    if is_neg:
        integer_part = integer_part[1:]
    thousands_list = []
    while len(integer_part) > 3:
        thousands_list.insert(0, integer_part[-3:])
        integer_part = integer_part[:-3]
    thousands_list.insert(0, integer_part)
    formatted_int = (('-' if is_neg else '') + thousands_sep.join(thousands_list))
    if precision > 0:
        res = f"{formatted_int}{dec_sep}{decimal_part}"
    else:
        res = formatted_int
    return _wrap_res(res, orig_path, enable_links)

@pass_context
def filter_percent(context, value, precision=2, force_symbol=True):
    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)
    val_raw = value.val if isinstance(value, TrackedValue) else value
    locale_code = get_locale_from_context(context)
    if val_raw in (None, ''):
        return _wrap_res('', orig_path, enable_links)
    
    if isinstance(val_raw, str):
        val_raw = val_raw.replace('%', '').strip()
        
    try:
        val = float(val_raw)
        precision = int(precision)
    except (ValueError, TypeError):
        return _wrap_res(val_raw, orig_path, enable_links)
    
    # Scale proportion to percentage if in range [-1.0, 1.0] non-zero (0.5 -> 50.0%)
    if -1.0 <= val <= 1.0 and val != 0:
        val = val * 100.0

    dec_sep = ','
    thousands_sep = '.'
    locale_lower = locale_code.lower()
    if any(x in locale_lower for x in ('en_us', 'en_gb', 'en_ca', 'en_au')):
        dec_sep = '.'
        thousands_sep = ','
        
    val_str = f"{val:.{precision}f}"
    parts = val_str.split('.')
    integer_part = parts[0]
    decimal_part = parts[1] if len(parts) > 1 else ''
    is_neg = integer_part.startswith('-')
    if is_neg:
        integer_part = integer_part[1:]
    thousands_list = []
    while len(integer_part) > 3:
        thousands_list.insert(0, integer_part[-3:])
        integer_part = integer_part[:-3]
    thousands_list.insert(0, integer_part)
    formatted_int = (('-' if is_neg else '') + thousands_sep.join(thousands_list))
    
    if precision > 0:
        res_num = f"{formatted_int}{dec_sep}{decimal_part}"
        if res_num.endswith(dec_sep + '00'):
            res_num = res_num[:-3]
        elif res_num.endswith('0') and dec_sep in res_num:
            res_num = res_num[:-1]
    else:
        res_num = formatted_int
        
    res = f"{res_num}%" if force_symbol else res_num
    return _wrap_res(res, orig_path, enable_links)

@pass_context
def filter_words(context, value, mode=None):
    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)
    val_raw = value.val if isinstance(value, TrackedValue) else value
    locale_code = get_locale_from_context(context)
    locale_lower = locale_code.lower()
    d = None
    if isinstance(val_raw, (datetime, date)):
        d = val_raw
    elif isinstance(val_raw, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                d = datetime.strptime(val_raw.strip(), fmt).date()
                break
            except ValueError:
                pass
    if d is not None:
        if 'es_es' in locale_lower:
            res = spanish_date_to_words(d)
        elif 'en_us' in locale_lower or 'en_gb' in locale_lower:
            res = english_date_to_words(d)
        else:
            res = catalan_date_to_words(d)
        return _wrap_res(res, orig_path, enable_links)
    try:
        val = float(val_raw)
    except (ValueError, TypeError):
        return _wrap_res(val_raw, orig_path, enable_links)
    if mode is None:
        mode = 'coin'
    if mode == 'coin':
        if 'es_es' in locale_lower:
            res = spanish_currency_to_words(val)
        elif 'en_us' in locale_lower or 'en_gb' in locale_lower:
            res = english_currency_to_words(val)
        else:
            res = catalan_currency_to_words(val)
    else:
        integer_part = int(val)
        decimal_part = int(round((val - integer_part) * 100))
        if 'es_es' in locale_lower:
            int_str = spanish_number_to_words(integer_part)
            if decimal_part > 0:
                res = f"{int_str} con {spanish_number_to_words(decimal_part)}"
            else:
                res = int_str
        elif 'en_us' in locale_lower or 'en_gb' in locale_lower:
            int_str = english_number_to_words(integer_part)
            if decimal_part > 0:
                res = f"{int_str} point {english_number_to_words(decimal_part)}"
            else:
                res = int_str
        else:
            int_str = catalan_number_to_words(integer_part)
            if decimal_part > 0:
                res = f"{int_str} amb {catalan_number_to_words(decimal_part)}"
            else:
                res = int_str
    return _wrap_res(res, orig_path, enable_links)

def filter_prefix(value, fallback, elided):
    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)
    val_raw = value.val if isinstance(value, TrackedValue) else value
    s = str(val_raw or '').strip()
    if not s:
        return _wrap_res(s, orig_path, enable_links)
    
    import re
    is_elision = bool(re.match(r"^(?:[aeiouàéèíóòúüïAEIOUÀÉÈÍÓÒÚÜÏ]|[hH][aeiouàéèíóòúüïAEIOUÀÉÈÍÓÒÚÜÏ])", s))
    pfx = elided if is_elision else fallback
    
    if pfx.endswith("'") or pfx.endswith("’"):
        res = f"{pfx}{s}"
    else:
        if pfx.endswith(" "):
            res = f"{pfx}{s}"
        else:
            res = f"{pfx} {s}"
    return _wrap_res(res, orig_path, enable_links)

class _DescKey:
    def __init__(self, obj):
        self.obj = obj
    def __lt__(self, other):
        return self.obj > other.obj
    def __gt__(self, other):
        return self.obj < other.obj
    def __eq__(self, other):
        return self.obj == other.obj
    def __le__(self, other):
        return self.obj >= other.obj
    def __ge__(self, other):
        return self.obj <= other.obj

def filter_sort(value, by=None, reverse=False, attribute=None, case_sensitive=False):
    if attribute and not by:
        by = attribute

    if value in (None, ''):
        return []

    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)

    if isinstance(value, (list, tuple)):
        lst = list(value)
    else:
        return value

    if not lst:
        return lst

    if by is None or by == '' or by == []:
        first_item = lst[0]
        if isinstance(first_item, dict) and len(first_item) > 0:
            keys = [k for k in first_item.keys() if not str(k).startswith('_')]
            if keys:
                by = [keys[0]]
            else:
                by = []
        else:
            def _scalar_key(item):
                v = item.val if hasattr(item, 'val') else item
                if v is None:
                    return (1, '')
                return (0, v)
            sorted_lst = sorted(lst, key=_scalar_key, reverse=reverse)
            return _wrap_tracked(sorted_lst, orig_path, enable_links) if orig_path else sorted_lst

    if isinstance(by, str):
        keys_spec = [by]
    elif isinstance(by, (list, tuple)):
        keys_spec = list(by)
    else:
        keys_spec = [str(by)]

    def _sort_key(item):
        raw_item = item
        keys_tuple = []
        for spec in keys_spec:
            spec_str = str(spec).strip()
            desc = False
            col_name = spec_str
            if spec_str.startswith('-'):
                desc = True
                col_name = spec_str[1:].strip()
            elif spec_str.startswith('+'):
                col_name = spec_str[1:].strip()

            val = None
            if isinstance(raw_item, dict):
                val = raw_item.get(col_name)
            elif hasattr(raw_item, col_name):
                val = getattr(raw_item, col_name)

            if hasattr(val, 'val'):
                val = val.val

            if val is None:
                sort_v = ''
                type_rank = 2
            elif isinstance(val, (int, float)):
                sort_v = -float(val) if desc else float(val)
                type_rank = 0
            else:
                s_v = str(val).lower() if not case_sensitive else str(val)
                sort_v = _DescKey(s_v) if desc else s_v
                type_rank = 1

            keys_tuple.append((type_rank, sort_v))
        return tuple(keys_tuple)

    sorted_lst = sorted(lst, key=_sort_key, reverse=reverse)
    return _wrap_tracked(sorted_lst, orig_path, enable_links) if orig_path else sorted_lst

def filter_where(value, criteria=None, **kwargs):
    if value in (None, ''):
        return []

    orig_path = getattr(value, '_path', None)
    enable_links = getattr(value, 'enable_links', True)

    if isinstance(value, (list, tuple)):
        lst = list(value)
    else:
        return value

    target_criteria = {}
    if isinstance(criteria, dict):
        target_criteria.update(criteria)
    elif isinstance(criteria, str) and kwargs:
        target_criteria[criteria] = list(kwargs.values())[0]
    target_criteria.update(kwargs)

    if not target_criteria:
        res = [item for item in lst if item]
        return _wrap_tracked(res, orig_path, enable_links) if orig_path else res

    def _matches_item(item):
        for col_name, req_val in target_criteria.items():
            col_str = str(col_name).strip()
            item_val = None
            if isinstance(item, dict):
                item_val = item.get(col_str)
            elif hasattr(item, col_str):
                item_val = getattr(item, col_str)

            if hasattr(item_val, 'val'):
                item_val = item_val.val

            def _single_match(act, exp):
                if exp is None:
                    return act is None or act == ''
                if act is None:
                    return False
                try:
                    act_num = float(act)
                    exp_num = float(exp)
                    return act_num == exp_num
                except (ValueError, TypeError):
                    pass
                return str(act).strip().lower() == str(exp).strip().lower()

            if isinstance(req_val, (list, tuple, set)):
                if not any(_single_match(item_val, rv) for rv in req_val):
                    return False
            else:
                if not _single_match(item_val, req_val):
                    return False
        return True

    res = [item for item in lst if _matches_item(item)]
    return _wrap_tracked(res, orig_path, enable_links) if orig_path else res

def is_excel_true(val):
    if hasattr(val, 'val'):
        val = val.val
    if val in (None, '', 0, 0.0, False, '0', '0.0'):
        return False
    if isinstance(val, str):
        v_str = val.strip().upper()
        if v_str in ('NO', 'FALS', 'FALSE', '0', '0.0', 'N', 'OFF', 'DESACTIVAT'):
            return False
    return True

def filter_cert(value):
    return is_excel_true(value)

def filter_fals(value):
    return not is_excel_true(value)

def _register_common_filters(env):
    """Wires elips' own Jinja2 filters/tests/globals onto an Environment.
    Shared by the two main render passes and by render_expression_preview
    so a filter behaves identically in the real document and in the
    variable modal's live preview."""
    env.filters['coin'] = filter_coin
    env.filters['number'] = filter_number
    env.filters['percent'] = filter_percent
    env.filters['percentatge'] = filter_percent
    env.filters['porcentaje'] = filter_percent
    env.filters['pct'] = filter_percent
    env.filters['words'] = filter_words
    env.filters['prefix'] = filter_prefix
    env.filters['sort'] = filter_sort
    env.filters['filter'] = filter_where
    env.filters['where'] = filter_where
    env.filters['CERT'] = filter_cert
    env.filters['cert'] = filter_cert
    env.filters['FALS'] = filter_fals
    env.filters['fals'] = filter_fals
    env.filters['IS_CERT'] = filter_cert
    env.filters['is_cert'] = filter_cert
    env.filters['IS_FALS'] = filter_fals
    env.filters['is_fals'] = filter_fals
    env.tests['CERT'] = filter_cert
    env.tests['cert'] = filter_cert
    env.tests['FALS'] = filter_fals
    env.tests['fals'] = filter_fals
    env.tests['is_cert'] = filter_cert
    env.tests['is_fals'] = filter_fals
    env.globals['TRUE'] = True
    env.globals['FALSE'] = False
    env.globals['true'] = True
    env.globals['false'] = False
    env.globals['CERT'] = filter_cert
    env.globals['FALS'] = filter_fals
    env.globals['cert'] = filter_cert
    env.globals['fals'] = filter_fals
    env.globals['is_excel_true'] = is_excel_true

def render_expression_preview(ctx_json, expr_str):
    """Evaluates a single Jinja2 expression (a variable path plus an
    optional filter chain, e.g. "pres.parts | sum(attribute='import')")
    against a JSON sample context — used by the template editor's variable
    modal to show a live preview of what a filter chain actually produces,
    without needing to render a whole document. Returns the real Python
    value (not stringified), so the caller can tell a scalar apart from a
    list/dict result. Never raises: any failure (bad path, wrong filter
    arity, an iterator not present in the sample context, ...) comes back
    as {success: False, error} instead, since a preview is best-effort by
    nature — the expression may reference loop iterators the caller could
    only approximate with a sample row.
    """
    try:
        ctx = json.loads(ctx_json)
        if not isinstance(ctx, dict):
            ctx = {}
        env = Environment(undefined=StrictUndefined, autoescape=False)
        _register_common_filters(env)
        compiled = env.compile_expression(expr_str, undefined_to_none=False)
        result = compiled(**ctx)
        return json.dumps({'success': True, 'result': result}, default=_custom_json_default, ensure_ascii=False)
    except Exception as ex:
        return json.dumps({'success': False, 'error': str(ex)}, ensure_ascii=False)

def render_json_text(excel_path, date_format='iso', strict=False):
    doc = excel_to_json(excel_path, date_format=date_format, strict=strict)
    return json.dumps(doc, ensure_ascii=False, default=_custom_json_default)

def _merge_live_json_into_doc(doc_val, live_val, depth=0, max_depth=20):
    """Recursively backfills doc_val (freshly re-parsed from the .xlsx) with
    anything present in live_val (the live in-browser JSON) that is missing
    or falsy on the doc side. Never removes or renames real data; only adds
    keys the Excel re-read couldn't have produced (e.g. calculated fields),
    at any nesting depth and inside tabular (list) rows."""
    if depth > max_depth:
        return doc_val
    if isinstance(doc_val, dict) and isinstance(live_val, dict):
        for k, v in live_val.items():
            if isinstance(k, str) and (k.startswith('_') or k in ('editor_metadata', 'editormetadata', '_hierarchy_schema', 'hierarchy_schema')):
                continue
            if k not in doc_val or not doc_val[k]:
                doc_val[k] = v
            else:
                doc_val[k] = _merge_live_json_into_doc(doc_val[k], v, depth + 1, max_depth)
        return doc_val
    if isinstance(doc_val, list) and isinstance(live_val, list):
        if len(doc_val) != len(live_val):
            # Row count diverged since the Excel was last regenerated (e.g. a
            # row was added/removed live): trust the live data wholesale.
            return live_val
        return [_merge_live_json_into_doc(d, l, depth + 1, max_depth) for d, l in zip(doc_val, live_val)]
    return doc_val

def _filter_empty_rows(data, visited=None, depth=0, max_depth=15):
    if depth > max_depth:
        return data
    if visited is None:
        visited = set()
    if isinstance(data, (dict, list)):
        data_id = id(data)
        if data_id in visited:
            return data
        visited.add(data_id)

    if isinstance(data, dict):
        new_dict = {}
        for k, v in data.items():
            if k in ('editor_metadata', '_hierarchy_schema'):
                new_dict[k] = v
            else:
                new_dict[k] = _filter_empty_rows(v, visited, depth + 1, max_depth)
        return new_dict
    elif isinstance(data, list):
        filtered_list = []
        for item in data:
            if isinstance(item, dict):
                if not _is_row_empty(item, visited.copy(), depth + 1, max_depth):
                    filtered_list.append(_filter_empty_rows(item, visited, depth + 1, max_depth))
            else:
                if not _is_value_empty(item):
                    filtered_list.append(_filter_empty_rows(item, visited, depth + 1, max_depth))
        return filtered_list
    return data

def _wrap_tracked(val, path='', enable_links=True, visited=None):
    if visited is None:
        visited = set()
    if isinstance(val, (dict, list)):
        val_id = id(val)
        if val_id in visited:
            return val
        visited.add(val_id)

    if isinstance(val, dict):
        if isinstance(val, TrackedDict):
            return val
        return TrackedDict(val, path, enable_links, visited)
    elif isinstance(val, list):
        if isinstance(val, TrackedList):
            return val
        return TrackedList(val, path, enable_links, visited)
    elif isinstance(val, (SafeDict, Placeholder)):
        return val
    elif isinstance(val, TrackedValue):
        return val
    return TrackedValue(val, path, enable_links)

class SafeDict(dict):
    def __init__(self, d, path='', visited=None):
        super().__init__()
        self._path = path
        if visited is None:
            visited = set()
        if isinstance(d, dict):
            for k, v in d.items():
                self[k] = _wrap_safe(v, f"{path}.{k}" if path else k, visited)

    def __getitem__(self, key):
        if key not in self:
            return Placeholder(f"{self._path}.{key}" if self._path else str(key))
        return super().__getitem__(key)

    def __getattr__(self, name):
        if name.startswith('_') or name in ('get', 'keys', 'items', 'values'):
            raise AttributeError(name)
        try:
            return self[name]
        except KeyError:
            raise AttributeError(name)

    def get(self, key, default=None):
        if key not in self:
            return Placeholder(f"{self._path}.{key}" if self._path else str(key))
        return super().get(key, default)

    def __str__(self):
        if '_default_val' in self:
            return str(self['_default_val'])
        return super().__str__()

def _wrap_safe(val, path='', visited=None):
    if visited is None:
        visited = set()
    if isinstance(val, (dict, list)):
        val_id = id(val)
        if val_id in visited:
            return val
        visited.add(val_id)

    if isinstance(val, dict):
        if isinstance(val, SafeDict):
            return val
        return SafeDict(val, path, visited)
    elif isinstance(val, list):
        return [_wrap_safe(item, f"{path}.{idx}", visited) for idx, item in enumerate(val)]
    elif isinstance(val, Placeholder):
        return val
    return val

def render_md_two_pass_with_report(excel_path, template_path, date_format='iso', strict=False):
    import traceback
    try:
        raw_doc = excel_to_json(excel_path, date_format=date_format, strict=strict)
        if isinstance(raw_doc, dict) and 'data' in raw_doc and isinstance(raw_doc['data'], dict):
            raw_data = raw_doc['data']
        else:
            raw_data = raw_doc
        doc = _filter_empty_rows(raw_data)

        # Foreign-key hydration (Select/dynamic fields, below) needs the
        # editor_metadata rows to know which fields are FKs and how to resolve
        # them — keep a reference to it *before* it gets popped off the tree
        # that Jinja2 actually sees. It may be replaced with a fresher copy
        # from /work/in.json below (see the live-state merge).
        fk_meta_list = doc.get('editor_metadata') or doc.get('editormetadata') or []

        # Remove internal metadata keys from main data model so Jinja2 context never sees them as data nodes
        doc.pop('_sheet_info', None)
        doc.pop('editor_metadata', None)
        doc.pop('_hierarchy_schema', None)
        doc.pop('editormetadata', None)
        doc.pop('hierarchy_schema', None)

        # Merge latest JSON from /work/in.json if present. This backfills anything
        # present in the live in-browser state (store.excelJsonData) but absent from
        # a fresh re-read of the .xlsx — most importantly, calculated/virtual fields
        # (editor_metadata type "Computed" or row-level CUSTOM formulas), which are
        # never written back as real Excel columns and so never survive excel_to_json
        # on their own. Recurses through nested groups and tabular (list) rows at any
        # depth, not just the top level, so calculated fields inside a `{% for %}`
        # loop over a sub-table are picked up too.
        try:
            if os.path.exists('/work/in.json'):
                with open('/work/in.json', 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    if isinstance(json_data, dict):
                        if 'data' in json_data and isinstance(json_data['data'], dict):
                            json_data = json_data['data']
                        live_meta = json_data.get('editor_metadata') or json_data.get('editorMetadata')
                        if isinstance(live_meta, list) and live_meta:
                            fk_meta_list = live_meta
                        for k, v in json_data.items():
                            if k.startswith('_') or k in ('editor_metadata', 'editormetadata', '_hierarchy_schema', 'hierarchy_schema'):
                                continue
                            if k not in doc or not doc[k]:
                                doc[k] = v
                            else:
                                doc[k] = _merge_live_json_into_doc(doc[k], v)
        except Exception:
            pass

        def _hydrate_foreign_keys(doc_dict, meta_list):
            if not isinstance(doc_dict, dict) or not isinstance(meta_list, list) or not meta_list:
                return doc_dict
            fk_map = {}
            for meta in meta_list:
                if isinstance(meta, dict):
                    m_type = str(meta.get('type', ''))
                    m_src = str(meta.get('sourceType', ''))
                    m_vec = str(meta.get('vectorPath', ''))
                    grp = str(meta.get('group', ''))
                    elem = str(meta.get('element', ''))
                    if m_type == 'Select' and m_src == 'dynamic' and m_vec and grp and elem:
                        fk_map[f"{grp}.{elem}"] = meta
                        short_grp = grp.split('.')[-1]
                        fk_map[f"{short_grp}.{elem}"] = meta
                        clean_grp = grp.replace('OUT_', '')
                        fk_map[f"{clean_grp}.{elem}"] = meta
                        clean_short = short_grp.replace('OUT_', '')
                        fk_map[f"{clean_short}.{elem}"] = meta
            if not fk_map:
                return doc_dict

            def _resolve_table(vec_path):
                if not vec_path:
                    return None
                if vec_path in doc_dict and isinstance(doc_dict[vec_path], list):
                    return doc_dict[vec_path]
                if f"OUT_{vec_path}" in doc_dict and isinstance(doc_dict[f"OUT_{vec_path}"], list):
                    return doc_dict[f"OUT_{vec_path}"]
                parts = vec_path.replace('doc.', '').replace('dades.', '').split('.')
                curr = doc_dict
                for p in parts:
                    if isinstance(curr, dict) and p in curr:
                        curr = curr[p]
                    else:
                        return None
                return curr if isinstance(curr, list) else None

            def _process_item(group_name, item):
                if isinstance(item, list):
                    for row in item:
                        _process_item(group_name, row)
                elif isinstance(item, dict):
                    for k, v in list(item.items()):
                        meta_key = f"{group_name}.{k}"
                        meta = None
                        if meta_key in fk_map:
                            meta = fk_map[meta_key]
                        else:
                            for fk_k, fk_m in fk_map.items():
                                if fk_k.endswith(f".{k}") or fk_k == k or (isinstance(fk_m, dict) and fk_m.get('element') == k):
                                    meta = fk_m
                                    break
                        if meta and v is not None and v != '' and not isinstance(v, (dict, list)):
                            tbl = _resolve_table(meta.get('vectorPath'))
                            if tbl and isinstance(tbl, list):
                                v_field = str(meta.get('valueField', ''))
                                d_field = str(meta.get('displayField', ''))
                                matched = None
                                for target_row in tbl:
                                    if isinstance(target_row, dict):
                                        v_k = v_field if (v_field and v_field in target_row) else (list(target_row.keys())[0] if target_row else '')
                                        d_k = d_field if (d_field and d_field in target_row) else v_k
                                        target_val = str(target_row.get(v_k, ''))
                                        target_disp = str(target_row.get(d_k, ''))
                                        if target_val == str(v) or target_disp == str(v):
                                            matched = target_row
                                            break
                                if matched:
                                    hydrated = dict(matched)
                                    hydrated['_default_val'] = v
                                    hydrated['value'] = v
                                    hydrated['val'] = v
                                    item[k] = hydrated
                        if isinstance(v, (dict, list)) and k not in ('editor_metadata', '_hierarchy_schema'):
                            child_path = f"{group_name}.{k}"
                            _process_item(child_path, v)

            for sheet_name, sheet_val in list(doc_dict.items()):
                if sheet_name not in ('editor_metadata', '_hierarchy_schema'):
                    _process_item(sheet_name, sheet_val)
            return doc_dict

        doc = _hydrate_foreign_keys(doc, fk_meta_list)

        with open(template_path, 'r', encoding='utf-8') as f:
            tpl_src = f.read()

        # Clean non-breaking spaces (\u00a0) that might be attached to Jinja2 tags
        tpl_src = tpl_src.replace('\u00a0', ' ')

        # See protect_inline_trailing_block_tags: keeps trim_blocks (below)
        # from eating row-separating newlines in TRANSPOSED_TABLE's inline
        # per-row {% for %}...{% endfor %} loops.
        tpl_src = protect_inline_trailing_block_tags(tpl_src)

        def _normalize_markdown_headings(text):
            if not text:
                return text
            lines = text.split('\\n')
            out = []
            for i, line in enumerate(lines):
                stripped = line.lstrip()
                if stripped.startswith('#') and i > 0:
                    if out and out[-1].strip() != '':
                        out.append('')
                out.append(line)
            return '\\n'.join(out)

        # Pass 1: Clean Context without HTML links (for Pandoc / Word export)
        clean_ctx = _wrap_safe(doc, '')
        if 'doc' not in clean_ctx:
            clean_ctx['doc'] = clean_ctx

        # trim_blocks/lstrip_blocks=True: a block tag's own trailing newline
        # (and, for lstrip_blocks, its leading indentation) is stripped
        # instead of ending up in the output — otherwise every
        # {% for %}/{% endfor %}/{% if %}/{% endif %} leaves a blank line
        # behind, which breaks any Markdown table generated by looping over
        # its rows (a blank line ends a table) and adds unwanted spacing
        # around conditionally-included paragraphs.
        env_clean = Environment(undefined=StrictUndefined, autoescape=False, trim_blocks=True, lstrip_blocks=True)
        _register_common_filters(env_clean)

        out1_clean, issues1 = render_with_recovery(env_clean, tpl_src, clean_ctx, 'primera')
        if '{{' in out1_clean or '{%' in out1_clean:
            out2_clean, issues2 = render_with_recovery(env_clean, out1_clean, clean_ctx, 'segona')
        else:
            out2_clean = out1_clean
            issues2 = []
        all_issues = issues1 + issues2

        # Pass 2: Tracked Context with HTML links (for HTML preview)
        html_ctx = _wrap_tracked(doc, '', enable_links=True)
        if 'doc' not in html_ctx:
            html_ctx['doc'] = html_ctx

        # trim_blocks/lstrip_blocks=True: see env_clean above.
        env_html = Environment(undefined=StrictUndefined, autoescape=False, trim_blocks=True, lstrip_blocks=True)
        _register_common_filters(env_html)

        out1_html, _ = render_with_recovery(env_html, tpl_src, html_ctx, 'primera_html')
        if '{{' in out1_html or '{%' in out1_html:
            out2_html, _ = render_with_recovery(env_html, out1_html, html_ctx, 'segona_html')
        else:
            out2_html = out1_html

        out2_clean = _normalize_markdown_headings(out2_clean)
        out2_html = _normalize_markdown_headings(out2_html)

        return json.dumps({
            'success': True,
            'markdown': out2_clean,
            'htmlMarkdown': out2_html,
            'issues': all_issues
        }, ensure_ascii=False, default=_custom_json_default)
    except Exception as ex:
        tb_str = traceback.format_exc()
        lineno = getattr(ex, 'lineno', None)
        msg = str(ex)
        return json.dumps({
            'success': False,
            'error': f"Error de conversió Jinja2: {msg}",
            'message': msg,
            'traceback': tb_str,
            'line': lineno
        }, ensure_ascii=False, default=_custom_json_default)


# =============================================================================
# Calculated-field evaluation engine (SI/ARRODONEIX/CONCAT/MONEDA/... mini-language)
# =============================================================================
# Ported from src/composables/useWasmEngines.js (evaluateCustomFormula /
# evaluateComputedFields), moved here so calculated-field formulas execute
# inside Pyodide's WASM sandbox instead of via a JS `new Function(...)`
# (dynamic eval) in the page's own execution context — the browser tab that
# also holds cookies, localStorage and DOM access. A hostile formula string
# smuggled in via editor_metadata now runs inside a restricted Python `eval`
# with no builtins beyond a small whitelist, isolated in the WASM sandbox with
# no path back to `window`/`document`/network APIs.
#
# Foreign-key hydration of dynamic Select fields (hydrateModelWithForeignKeys
# in useWasmEngines.js) stays in JS: it is a data-shape transformation, not
# calculation, and JS calls it before handing data to evaluate_computed_fields.

_CEF_RESERVED_TOKENS = {
    'SI', 'IF', 'ARRODONEIX', 'ROUND', 'ABS', 'MIN', 'MAX', 'OR', 'O', 'AND', 'I',
    'ANY', 'SOME', 'EVERY', 'ALL', 'Math', '__round', '__or', '__and',
    'CERT', 'FALS', 'cert', 'fals', 'is_cert', 'is_fals', '__is_cert', '__is_fals',
    'true', 'false', 'null', 'undefined', 'doc', 'dades', 'return', 'function',
    'abs', 'True', 'False', 'None', 'and', 'or', 'not', 'if', 'else', 'is', 'in',
}

_CEF_TOKEN_RE = re.compile(
    r'\b(?:[a-zA-Z_][a-zA-Z0-9_]*|doc\.[a-zA-Z0-9_.]+|dades\.[a-zA-Z0-9_.]+)'
    r'(?:\[\d+\])?(?:\.[a-zA-Z_][a-zA-Z0-9_.]*(?:\[\d+\])?)*\b'
)
_CEF_ARRAY_INDEX_RE = re.compile(r'^([a-zA-Z0-9_]+)\[(\d+)\]$')


def _cef_balanced_call_args(s, open_paren_idx):
    """Finds the matching ')' for a '(' at open_paren_idx, respecting nested
    parens and quoted strings. Returns (end_idx, args_str) or (None, None)."""
    depth = 1
    in_quotes = False
    quote_char = ''
    end_idx = None
    for i in range(open_paren_idx + 1, len(s)):
        ch = s[i]
        if in_quotes:
            if ch == quote_char:
                in_quotes = False
        elif ch in ('"', "'"):
            in_quotes = True
            quote_char = ch
        elif ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
            if depth == 0:
                end_idx = i
                break
    if end_idx is None:
        return None, None
    return end_idx, s[open_paren_idx + 1:end_idx]


def _cef_split_args(args_str):
    """Splits a function-call argument string on ';' or ',' at paren/quote depth 0."""
    parts = []
    current = ''
    depth = 0
    in_quotes = False
    quote_char = ''
    for ch in args_str:
        if in_quotes:
            current += ch
            if ch == quote_char:
                in_quotes = False
        elif ch in ('"', "'"):
            in_quotes = True
            quote_char = ch
            current += ch
        elif ch == '(':
            depth += 1
            current += ch
        elif ch == ')':
            depth -= 1
            current += ch
        elif ch in (';', ',') and depth == 0:
            parts.append(current.strip())
            current = ''
        else:
            current += ch
    parts.append(current.strip())
    return parts


def _cef_transform_if(s):
    """SI(cond; a; b) / IF(cond; a; b) -> Python ternary: ((a) if (cond) else (b))."""
    prev = None
    while prev != s:
        prev = s
        m = re.search(r'\b(SI|IF)\s*\(', s, re.IGNORECASE)
        if not m:
            break
        end_idx, args_str = _cef_balanced_call_args(s, m.end() - 1)
        if end_idx is None:
            break
        full_match = s[m.start():end_idx + 1]
        parts = _cef_split_args(args_str)
        if len(parts) >= 3:
            cond, t_val = parts[0], parts[1]
            f_val = ';'.join(parts[2:])
            s = s.replace(full_match, f'(({t_val}) if ({cond}) else ({f_val}))', 1)
        else:
            break
    return s


def _cef_transform_round(s):
    """ARRODONEIX(valor; decimals) / ROUND(...) -> __round(valor, decimals)."""
    prev = None
    while prev != s:
        prev = s
        m = re.search(r'\b(ARRODONEIX|ROUND)\s*\(', s, re.IGNORECASE)
        if not m:
            break
        end_idx, args_str = _cef_balanced_call_args(s, m.end() - 1)
        if end_idx is None:
            break
        full_match = s[m.start():end_idx + 1]
        parts = _cef_split_args(args_str)
        val_expr = parts[0] if parts and parts[0] else '0'
        prec_expr = parts[1] if len(parts) > 1 else '0'
        s = s.replace(full_match, f'__round({val_expr}, {prec_expr})', 1)
    return s


def _cef_transform_or_and(s):
    """OR(path)/AND(path) (plus O/I/ANY/SOME/EVERY/ALL synonyms) -> __or("path")/__and("path")."""
    prev = None
    while prev != s:
        prev = s
        m = re.search(r'\b(OR|O|AND|I|ANY|SOME|EVERY|ALL)\s*\(', s, re.IGNORECASE)
        if not m:
            break
        end_idx, arg_str = _cef_balanced_call_args(s, m.end() - 1)
        if end_idx is None:
            break
        fn_name = m.group(1).upper()
        full_match = s[m.start():end_idx + 1]
        # Escape backslashes before quotes (the original JS port only escaped
        # quotes, which could desync the generated string literal).
        escaped = arg_str.strip().replace('\\', '\\\\').replace('"', '\\"')
        if fn_name in ('OR', 'O', 'ANY', 'SOME'):
            s = s.replace(full_match, f'__or("{escaped}")', 1)
        elif fn_name in ('AND', 'I', 'EVERY', 'ALL'):
            s = s.replace(full_match, f'__and("{escaped}")', 1)
        else:
            break
    return s


def _cef_transform_cert_fals(s):
    s = re.sub(r'\b(CERT|is_cert)\s*\(', '__is_cert(', s, flags=re.IGNORECASE)
    s = re.sub(r'\b(FALS|is_fals)\s*\(', '__is_fals(', s, flags=re.IGNORECASE)
    return s


def _cef_parse_num_or_string(raw_val):
    """Returns either a number/bool (embedded later as a literal) or a Python
    `str` that is ALREADY valid quoted Python source (via repr) — mirroring
    the JS version's dual return type, which the caller dispatches on."""
    if isinstance(raw_val, bool):
        return raw_val
    if isinstance(raw_val, (int, float)):
        return raw_val
    if isinstance(raw_val, str):
        if raw_val.strip() == '':
            return 0
        try:
            return float(raw_val.replace(',', '.'))
        except ValueError:
            return repr(raw_val)
    return 0


def _cef_get_nested_value(obj, parts, global_data):
    current = obj
    for part in parts:
        if current is None:
            return None
        m = _CEF_ARRAY_INDEX_RE.match(part)
        if m:
            arr_key, idx = m.group(1), int(m.group(2))
            arr = current.get(arr_key) if isinstance(current, dict) else None
            current = arr[idx] if isinstance(arr, list) and idx < len(arr) else None
            if current is None:
                return None
        elif isinstance(current, list):
            nums = []
            for item in current:
                if isinstance(item, dict) and part in item:
                    try:
                        nums.append(float(item[part]))
                    except (TypeError, ValueError):
                        pass
            return sum(nums) if nums else 0
        elif isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, (str, int, float)):
            # Fallback: current is a scalar FK value (e.g. "PART-01"); look it
            # up in the global tables for a row that carries this `part` field.
            found_val = None
            if global_data:
                for table in global_data.values():
                    if isinstance(table, list):
                        match_row = next((r for r in table if isinstance(r, dict)
                                           and any(str(v) == str(current) for v in r.values())), None)
                        if match_row is not None and part in match_row:
                            found_val = match_row[part]
                            break
            if found_val is not None:
                current = found_val
            else:
                return None
        else:
            return None
    return current


def _cef_resolve_value(path_str, row, global_data):
    if not path_str:
        return None
    if isinstance(row, dict) and path_str in row:
        return _cef_parse_num_or_string(row[path_str])
    clean_path = re.sub(r'^(doc|dades)\.', '', path_str, flags=re.IGNORECASE)
    path_parts = [p for p in clean_path.split('.') if p]

    val = _cef_get_nested_value(row, path_parts, global_data)
    if val is not None:
        return _cef_parse_num_or_string(val)

    if global_data:
        val = _cef_get_nested_value(global_data, path_parts, global_data)
        if val is None and path_parts:
            prefixed = ['OUT_' + path_parts[0]] + path_parts[1:]
            val = _cef_get_nested_value(global_data, prefixed, global_data)
        if val is not None:
            return _cef_parse_num_or_string(val)
    return None


def _cef_is_cert(val):
    if val in (None, False, '', 0, 0.0, '0', '0.0'):
        return False
    if isinstance(val, str) and val.strip().upper() in ('NO', 'FALS', 'FALSE', '0', '0.0', 'N', 'OFF', 'DESACTIVAT'):
        return False
    return True


def _cef_is_fals(val):
    return not _cef_is_cert(val)


def _cef_extract_bool_val(val):
    if val in (None, False, 0, '0', '', '0.0'):
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    if isinstance(val, str):
        return val.strip().lower() in ('true', '1', 'si', 'sí', 'cert', 'yes')
    return bool(val)


def _cef_eval_or_and(path_str, mode, row, global_data):
    clean_path = re.sub(r"^['\"]|['\"]$", '', path_str.strip())
    clean_path = re.sub(r'^(doc|dades)\.', '', clean_path, flags=re.IGNORECASE)
    parts = [p for p in clean_path.split('.') if p]
    if not parts:
        return False

    def collect_values(start_obj):
        if not isinstance(start_obj, (dict, list)):
            return None
        current = start_obj
        for i, part in enumerate(parts):
            if current is None:
                return None
            if isinstance(current, list):
                remaining = parts[i:]
                out = []
                for item in current:
                    if not isinstance(item, dict):
                        out.append(item)
                        continue
                    sub = item
                    for p in remaining:
                        if sub is None:
                            sub = None
                            break
                        sub = sub.get(p) if isinstance(sub, dict) else None
                    out.append(sub)
                return out
            current = current.get(part) if isinstance(current, dict) else None
        return current if isinstance(current, list) else [current]

    lst = collect_values(row)
    if (not lst) and global_data:
        lst = collect_values(global_data)
        if (not lst) and parts:
            sheet_key = parts[0]
            prefixed = global_data.get('OUT_' + sheet_key)
            if prefixed is not None:
                lst = collect_values({'OUT_' + sheet_key: prefixed})

    if not isinstance(lst, list):
        lst = []

    if mode == 'OR':
        return any(_cef_extract_bool_val(v) for v in lst)
    return len(lst) > 0 and all(_cef_extract_bool_val(v) for v in lst)


def _cef_normalize_number(n):
    """Rounds to 6 decimals and collapses whole-number floats to int (e.g.
    30.0 -> 30). JS's single Number type never carries this distinction —
    JSON.stringify(30.0) is "30" — so without this, a value that used to
    render as "30" in a generated document would render as "30.0" now that
    calculation happens in Python."""
    rounded = round(float(n) * 1000000) / 1000000
    if rounded == int(rounded) and abs(rounded) < 1e15:
        return int(rounded)
    return rounded


def evaluate_custom_formula(formula_str, row, global_data=None):
    """Evaluates one CUSTOM-formula string (the SI/ARRODONEIX/CONCAT/MONEDA/...
    mini-language) against `row` (the record being computed) and `global_data`
    (the whole data tree, for cross-group lookups). Returns a number, bool or
    string, or `row.get(formula_str)` as a last-resort fallback on any error —
    matching the JS version's forgiving behavior (a broken formula shouldn't
    crash the form, just leave the field showing something recognizable)."""
    if not formula_str or not isinstance(formula_str, str):
        return 0
    try:
        expr = formula_str.strip()

        # Defense in depth: even with a restricted __builtins__, Python's eval()
        # still allows plain attribute access, which is enough for the classic
        # `().__class__.__bases__[0].__subclasses__()`-style sandbox escape (that
        # chain, and `__import__(...)`, both require a literal '__'). A bare
        # `import`/`eval`/`open`/... call can't do anything here regardless —
        # eval() only accepts expressions, so `import` is a SyntaxError inside
        # it, and every other name in that family is simply absent from the
        # restricted __builtins__ below — so only '__' itself needs blocking.
        # (Field names like "import" — Catalan for "amount" — are common and
        # legitimate in this app's domain and must not be caught here.)
        if '__' in expr:
            return row.get(formula_str, 0) if isinstance(row, dict) else 0

        expr = _cef_transform_if(expr)
        expr = _cef_transform_round(expr)
        expr = _cef_transform_or_and(expr)
        expr = _cef_transform_cert_fals(expr)

        expr = re.sub(r'\bABS\s*\(', 'abs(', expr, flags=re.IGNORECASE)
        expr = re.sub(r'(^|[^<>=!])=([^=])', r'\1==\2', expr)
        expr = expr.replace('<>', '!=')
        expr = expr.replace('^', '**')

        found_tokens = set()
        for m in _CEF_TOKEN_RE.finditer(expr):
            t = m.group(0)
            if t not in _CEF_RESERVED_TOKENS and t.upper() not in _CEF_RESERVED_TOKENS:
                found_tokens.add(t)

        for t in sorted(found_tokens, key=len, reverse=True):
            resolved = _cef_resolve_value(t, row, global_data)
            if resolved is not None:
                replacement = resolved if isinstance(resolved, str) else f'({resolved!r})'
                expr = re.sub(r'\b' + re.escape(t) + r'\b', lambda _m: replacement, expr)

        safe_globals = {
            '__builtins__': {'abs': abs, 'min': min, 'max': max, 'True': True, 'False': False, 'None': None},
            '__round': lambda val, prec=0: round(float(val), int(prec)) if str(val).strip() != '' else 0,
            '__is_cert': _cef_is_cert,
            '__is_fals': _cef_is_fals,
            '__or': lambda arg: _cef_eval_or_and(str(arg), 'OR', row, global_data),
            '__and': lambda arg: _cef_eval_or_and(str(arg), 'AND', row, global_data),
            'CERT': _cef_is_cert, 'FALS': _cef_is_fals, 'cert': _cef_is_cert, 'fals': _cef_is_fals,
        }
        result = eval(expr, safe_globals, {})

        if isinstance(result, bool):
            return result
        if isinstance(result, (int, float)):
            return _cef_normalize_number(result)
        if result is not None:
            return str(result)
        return 0
    except Exception:
        return row.get(formula_str, 0) if isinstance(row, dict) else 0


def _cef_is_custom_fn(fn, formula):
    if formula and str(formula).strip():
        return True
    upper = (fn or '').upper()
    return upper in ('CUSTOM', 'FORMULA', '', 'NONE')


def _cef_is_group_match(meta_group, hint):
    if not meta_group:
        return True
    if not hint:
        return False
    clean_m = re.sub(r'^OUT_', '', meta_group, flags=re.IGNORECASE).lower()
    clean_h = re.sub(r'^OUT_', '', hint, flags=re.IGNORECASE).lower()
    if clean_m == clean_h:
        return True
    if clean_m.split('.')[-1] == clean_h.split('.')[-1]:
        return True
    root_hints = {'doc', 'dades', 'global', 'header', 'general', 'presupost', 'pressupost', 'resum', 'summary', 'root', 'main', ''}
    return clean_h in root_hints and clean_m in root_hints


def _cef_extract_val(child, col):
    if child is None:
        return 0
    if isinstance(child, dict):
        if col and child.get(col) is not None:
            try:
                return float(child[col])
            except (TypeError, ValueError):
                return 0
        for k, v in child.items():
            if not str(k).startswith('_'):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    continue
        return 0
    try:
        return float(child)
    except (TypeError, ValueError):
        return 0


def evaluate_computed_fields(data_json, metadata_json, debug_mode=False):
    """Two-phase bottom-up evaluation of every calculated field in `data_json`:
    first row-level CUSTOM formulas (evaluate_custom_formula), then SUM/COUNT/
    AVERAGE/MIN/MAX/OR/AND aggregations that read the just-computed CUSTOM
    values from child tables. Mutates and returns a JSON string of the full
    data tree (mirroring the JS version's in-place mutation of store.excelJsonData,
    so Vue's reactivity still targets the same paths after the round trip).

    `data_json` must already have dynamic Select fields hydrated (done in JS
    by hydrateModelWithForeignKeys before this is called) — this function only
    computes formulas/aggregations, it does not resolve foreign keys.
    """
    try:
        data = json.loads(data_json)
        metadata = json.loads(metadata_json) if metadata_json else []
    except Exception as ex:
        return json.dumps({'success': False, 'error': str(ex)})

    if not isinstance(data, dict):
        return json.dumps({'success': True, 'data': data, 'logs': []})

    logs = []

    computed_metas = [m for m in metadata if isinstance(m, dict) and (
        m.get('isCalculated') is True or
        m.get('type') == 'Computed' or
        m.get('sourceType') == 'computed' or
        (m.get('calcFn') and m.get('calcFn') not in ('', 'NONE')) or
        (m.get('calcFormula') and str(m.get('calcFormula')).strip() != '')
    )]

    if debug_mode:
        if not computed_metas:
            logs.append(f"🧮 [DEPURACIÓ CAMPS CALCULATS] No s'ha trobat cap camp marcat com a calculat (editor_metadata té {len(metadata)} metadades en total).")
        else:
            details = '\n'.join(
                f"  • [Grup: {m.get('group') or 'global'} | Camp: {m.get('element')}] "
                f"Tipus: {m.get('calcFn') or 'CUSTOM'} | Fórmula/Vector: \"{m.get('calcFormula') or m.get('calcVector') or ''}\""
                for m in computed_metas
            )
            logs.append(f"🧮 [DEPURACIÓ CAMPS CALCULATS] Detectats {len(computed_metas)} camps calculats:\n{details}")

    if not computed_metas:
        return json.dumps({'success': True, 'data': data, 'logs': logs})

    custom_metas = [m for m in computed_metas if _cef_is_custom_fn(m.get('calcFn'), m.get('calcFormula')) and m.get('calcFormula')]
    agg_metas = [m for m in computed_metas if not _cef_is_custom_fn(m.get('calcFn'), m.get('calcFormula'))]

    def run_custom_pass(container, group_hint='', visited=None):
        if visited is None:
            visited = set()
        if not isinstance(container, (dict, list)):
            return
        cid = id(container)
        if cid in visited:
            return
        visited.add(cid)

        if isinstance(container, list):
            for item in container:
                run_custom_pass(item, group_hint, visited)
            return

        for k, v in container.items():
            if k not in ('_sheet_info', '_hierarchy_schema', 'editor_metadata') and isinstance(v, (dict, list)):
                run_custom_pass(v, k if isinstance(v, list) else group_hint, visited)

        for meta in custom_metas:
            if _cef_is_group_match(meta.get('group'), group_hint):
                calculated_val = evaluate_custom_formula(meta.get('calcFormula'), container, data)
                if calculated_val is not None:
                    old_val = container.get(meta.get('element'))
                    container[meta.get('element')] = calculated_val
                    if debug_mode:
                        logs.append(f"✨ [CÀLCUL CUSTOM] {meta.get('group') or group_hint}.{meta.get('element')} = {calculated_val} "
                                    f"(Fórmula: \"{meta.get('calcFormula')}\", Anterior: {old_val})")

    def find_sub_list(obj, target_vec, sub_visited=None, depth=0):
        """Recursively searches container (dict values or list items, at any
        depth up to 5) for the first dict that has target_vec as an array key."""
        if sub_visited is None:
            sub_visited = set()
        if not isinstance(obj, (dict, list)) or depth > 5:
            return None
        oid = id(obj)
        if oid in sub_visited:
            return None
        sub_visited.add(oid)

        if isinstance(obj, dict) and isinstance(obj.get(target_vec), list):
            return obj[target_vec]

        children = obj.values() if isinstance(obj, dict) else obj
        for child in children:
            if isinstance(child, (dict, list)):
                found = find_sub_list(child, target_vec, sub_visited, depth + 1)
                if found is not None:
                    return found
        return None

    def run_agg_pass(container, group_hint='', visited=None):
        if visited is None:
            visited = set()
        if not isinstance(container, (dict, list)):
            return
        cid = id(container)
        if cid in visited:
            return
        visited.add(cid)

        if isinstance(container, list):
            for item in container:
                run_agg_pass(item, group_hint, visited)
            return

        for k, v in container.items():
            if k not in ('_sheet_info', '_hierarchy_schema', 'editor_metadata') and isinstance(v, (dict, list)):
                run_agg_pass(v, k if isinstance(v, list) else group_hint, visited)

        for meta in agg_metas:
            target_vec = meta.get('calcVector')
            fn = (meta.get('calcFn') or 'SUM').upper()
            col = meta.get('calcTargetCol')

            if _cef_is_group_match(meta.get('group'), group_hint) or (target_vec and isinstance(container.get(target_vec), list)):
                child_list = None
                if target_vec and isinstance(container.get(target_vec), list):
                    child_list = container[target_vec]
                elif target_vec and isinstance(data.get(target_vec), list):
                    child_list = data[target_vec]
                elif target_vec:
                    child_list = find_sub_list(container, target_vec)

                if child_list is not None:
                    if fn == 'COUNT':
                        calculated_val = len(child_list)
                    elif fn == 'SUM':
                        calculated_val = _cef_normalize_number(sum(_cef_extract_val(c, col) for c in child_list))
                    elif fn in ('AVG', 'AVERAGE'):
                        nums = [_cef_extract_val(c, col) for c in child_list]
                        calculated_val = _cef_normalize_number(sum(nums) / len(nums) if nums else 0)
                    elif fn == 'MIN':
                        nums = [_cef_extract_val(c, col) for c in child_list]
                        calculated_val = _cef_normalize_number(min(nums)) if nums else 0
                    elif fn == 'MAX':
                        nums = [_cef_extract_val(c, col) for c in child_list]
                        calculated_val = _cef_normalize_number(max(nums)) if nums else 0
                    elif fn in ('OR', 'O', 'SOME', 'ANY'):
                        calculated_val = any(_cef_extract_bool_val(c.get(col) if col and isinstance(c, dict) else c) for c in child_list)
                    elif fn in ('AND', 'I', 'EVERY', 'ALL'):
                        calculated_val = len(child_list) > 0 and all(_cef_extract_bool_val(c.get(col) if col and isinstance(c, dict) else c) for c in child_list)
                    else:
                        continue

                    old_val = container.get(meta.get('element'))
                    container[meta.get('element')] = calculated_val
                    if debug_mode:
                        logs.append(f"📊 [CÀLCUL AGREGACIÓ] {meta.get('group') or group_hint}.{meta.get('element')} = {calculated_val} "
                                    f"({fn} de '{target_vec}' [{len(child_list)} elements], Anterior: {old_val})")
                elif debug_mode:
                    logs.append(f"⚠️ [CÀLCUL AGREGACIÓ] No s'ha trobat la llista '{target_vec}' per calcular {meta.get('element')}.")

    for key, val in list(data.items()):
        if key not in ('_sheet_info', '_hierarchy_schema', 'editor_metadata'):
            run_custom_pass(val, key)

    for key, val in list(data.items()):
        if key not in ('_sheet_info', '_hierarchy_schema', 'editor_metadata'):
            run_agg_pass(val, key)

    return json.dumps({'success': True, 'data': data, 'logs': logs}, ensure_ascii=False, default=_custom_json_default)
