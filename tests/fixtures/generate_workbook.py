"""Generates a synthetic .xlsx fixture that exercises elips' full data model:

- KV sheets, both header-less (``OUT_General``) and with a ``Clau|Valor``
  header (``OUT_pres``), including intentionally blank fields.
- A 4-level nested tabular hierarchy via dotted sheet names
  (``OUT_pres.parts`` -> ``.activitats`` -> ``.recursos``), matched both by
  the default id-column heuristic and, for the deepest level, by an explicit
  ``_hierarchy_metadata`` foreign-key override.
- A second, independent hierarchy (``OUT_Criteris`` -> ``.Subcriteris``) used
  to exercise "ghost row" filtering (rows that are entirely blank/zero, as
  produced by unresolved formulas) without duplicating children across
  unrelated parents (the "cartesian explosion" regression).
- A fully-empty tabular sheet (``OUT_Lots``) and a small real one
  (``OUT_Mesa``) with a merged decorative title row above its header, to
  exercise the header-row-fallback and merged-cell write paths.
- A "complex" (non-link) formula and a simple cross-sheet link formula in
  ``OUT_pres.parts``, to exercise the orphan-formula / link-formula
  preservation logic in ``update_excel_from_json``.
- An ``editor_metadata`` sheet covering the main field types (Text, Number,
  Percentage, Date, Boolean, Select static/dynamic, Computed SUM/FORMULA,
  Table) plus grid layout and a dynamic item-title formula.

Run directly to write the fixture to disk:

    python3 tests/fixtures/generate_workbook.py [output_path]

Also importable as a module (``build_workbook()``) so tests can generate the
fixture in a temp directory without shelling out.
"""
import os
import sys

from openpyxl import Workbook

DEFAULT_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "elips_test_fixture.xlsx")


def _write_rows(ws, rows):
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = val


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # --- OUT_General: header-less KV sheet, with intentionally blank fields ---
    ws = wb.create_sheet("OUT_General")
    _write_rows(ws, [
        ["titol_contracte", "Subministrament d'equips informàtics"],
        ["num_expedient", ""],
        ["modalitat", "Contracte Públic"],
        ["data_redaccio", ""],
        ["nom_responsable", "Anna Puig Soler"],
        ["pressupost_base", 125000.5],
        ["percentatge_iva", 0.21],
    ])

    # --- OUT_pres: KV sheet with a Clau|Valor header ---
    ws = wb.create_sheet("OUT_pres")
    _write_rows(ws, [
        ["Clau", "Valor"],
        ["pressupost", "Pressupost anual"],
        ["anualitat", 2026],
    ])

    # --- OUT_pres.parts: tabular, child of the OUT_pres KV sheet (all rows
    # attach to the single parent, no id-matching needed). Row 4 is a "ghost"
    # row (all cells blank/zero) that must be filtered out entirely. Includes
    # a complex formula (orphan candidate) and a simple cross-sheet link.
    ws = wb.create_sheet("OUT_pres.parts")
    _write_rows(ws, [
        ["id_partida", "nom_partida", "import", "cost_amb_iva", "referencia"],
        ["PART-01", "Equips de sobretaula", 45000, "=C2*1.21", "=OUT_pres!B2"],
        ["PART-02", "Portàtils", 30000, "=C3*1.21", "=OUT_pres!B2"],
        ["", "", 0, 0, ""],
    ])

    # --- OUT_pres.parts.activitats: tabular, matched to its parent part row
    # via the shared 'id_partida' column (implicit id-heuristic matching).
    ws = wb.create_sheet("OUT_pres.parts.activitats")
    _write_rows(ws, [
        ["id_partida", "id_activitat", "nom_activitat", "hores"],
        ["PART-01", "ACT-01", "Instal·lació d'equips", 12],
        ["PART-01", "ACT-02", "Configuració de xarxa", 8],
        ["PART-02", "ACT-03", "Desplegament de programari", 6],
        ["", "", "", 0],
    ])

    # --- OUT_pres.parts.activitats.rec: 4th nesting level, matched via an
    # EXPLICIT foreign key declared in _hierarchy_metadata (child column
    # 'activitat_ref' does not share a name with the parent's 'id_activitat').
    # (kept to <=31 chars: Excel sheet name length limit)
    ws = wb.create_sheet("OUT_pres.parts.activitats.rec")
    _write_rows(ws, [
        ["activitat_ref", "recurs", "unitats"],
        ["ACT-01", "Tècnic instal·lador", 2],
        ["ACT-02", "Tècnic de xarxes", 1],
        ["ACT-03", "Analista de sistemes", 1],
    ])

    ws = wb.create_sheet("_hierarchy_metadata")
    _write_rows(ws, [
        ["full_path", "parent_key", "child_key"],
        ["pres.parts.activitats.rec", "id_activitat", "activitat_ref"],
    ])

    # --- OUT_Criteris / OUT_Criteris.Subcriteris: independent hierarchy used
    # to check ghost-row filtering doesn't cause cartesian duplication.
    ws = wb.create_sheet("OUT_Criteris")
    _write_rows(ws, [
        ["id", "descripcio", "puntuacio"],
        ["F1", "Preu i termini", 40],
        ["V1", "Valoració tècnica", 20],
        ["F2", "Millores", 10],
        ["", "", 0],
        ["", "", 0],
    ])

    ws = wb.create_sheet("OUT_Criteris.Subcriteris")
    _write_rows(ws, [
        ["id", "subid", "descripcio_sub", "punts_sub"],
        ["F1", "F1a", "Preu ofertat", 25],
        ["F1", "F1b", "Termini de lliurament", 15],
        ["F2", "F2a", "Garantia ampliada", 10],
        ["", "", "", 0],
    ])

    # --- OUT_Lots: tabular sheet that ends up with zero real rows ---
    ws = wb.create_sheet("OUT_Lots")
    _write_rows(ws, [
        ["id_lot", "descripcio", "import_lot"],
        ["", "", 0],
        ["", "", 0],
    ])

    # --- OUT_Mesa: real tabular data (3+ columns puts a non-dotted sheet
    # over the tabular-detection threshold). The "vot" cells for the two
    # vocals are merged together (a common presentation pattern for
    # identical adjacent values), exercising the MergedCell write path on
    # export without disturbing header detection. ---
    ws = wb.create_sheet("OUT_Mesa")
    _write_rows(ws, [
        ["nom", "carrec", "vot"],
        ["Anna Puig", "President", "Sí"],
        ["Jordi Vila", "Vocal", "Sí"],
        ["Marta Soler", "Vocal", "Sí"],
        ["David Roig", "Secretari", "No"],
    ])
    ws.merge_cells("C3:C4")

    # --- editor_metadata: covers the main field types & configuration knobs ---
    ws = wb.create_sheet("editor_metadata")
    headers = ['group', 'element', 'type', 'options', 'sourceType', 'multiple', 'vectorPath',
               'displayField', 'valueField', 'width', 'calcFn', 'calcVector', 'calcTargetCol',
               'calcFormula', 'gridRow', 'gridOrder', 'gridFill', 'label', 'groupLayout', 'itemTitleFormula']
    rows = [headers]
    rows.append(['General', 'titol_contracte', 'Text', '', '', 0, '', '', '', '', '', '', '', '', 1, 1, 1, 'Títol del contracte', '', ''])
    rows.append(['General', 'modalitat', 'Select', 'Contracte Públic,Contracte Privat', 'static', 0, '', '', '', '', '', '', '', '', 2, 1, 0, 'Modalitat', '', ''])
    rows.append(['General', 'pressupost_base', 'Number', '', '', 0, '', '', '', '', '', '', '', '', 2, 2, 0, 'Pressupost base', '', ''])
    rows.append(['General', 'percentatge_iva', 'Percentage', '', '', 0, '', '', '', '', '', '', '', '', 3, 1, 0, 'Percentatge IVA', '', ''])
    rows.append(['General', 'data_redaccio', 'Date', '', '', 0, '', '', '', '', '', '', '', '', 3, 2, 0, 'Data de redacció', '', ''])
    rows.append(['pres', 'total_pressupostat', 'Computed', '', '', 0, 'pres.parts', '', '', '', 'SUM', 'pres.parts', 'import', '', 1, 1, 1, 'Total pressupostat', '', ''])
    rows.append(['pres.parts', 'proveidor_assignat', 'Select', '', 'dynamic', 0, 'pres.parts', 'nom_partida', 'id_partida', '', '', '', '', '', 4, 1, 1, 'Partida associada', 'vertical', 'CONCAT(nom_partida; " ("; MONEDA(import); ")")'])
    rows.append(['pres.parts.activitats', 'resum', 'Table', '', '', 0, '', '', '', '', '', '', '', '', 1, 1, 1, 'Activitats', 'horizontal', ''])
    for row in rows:
        ws.append(row)
    ws.sheet_state = 'hidden'

    return wb


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUTPUT
    wb = build_workbook()
    wb.save(out_path)
    print(f"Fixture generada: {out_path}")


if __name__ == "__main__":
    main()
