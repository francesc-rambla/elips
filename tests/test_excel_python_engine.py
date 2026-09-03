# elips — Editor de LIcitacions PúbliqueS
# Copyright (C) 2026  Francesc Rambla i Marigot
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""Tests the Python data engine (src/python/engine.py) that elips runs inside
Pyodide, driving it through a synthetic fixture workbook generated on the fly
by tests/fixtures/generate_workbook.py rather than an external, untracked
.xlsx file. The fixture exercises the app's documented feature set: header-less
and headered KV sheets, a 4-level nested tabular hierarchy (matched both by
the implicit id-column heuristic and by an explicit _hierarchy_metadata
foreign key), ghost-row filtering without cartesian explosion across
unrelated parents, ghost/merged-cell round-trip export, complex-formula
orphan preservation, and the two-pass Jinja2 rendering pipeline with several
custom filters.
"""
import os
import sys
import json
import shutil
import tempfile
import unittest
import importlib.util
from unittest import mock

from openpyxl import load_workbook

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIXTURES_DIR = os.path.join(REPO_ROOT, "tests", "fixtures")
sys.path.insert(0, FIXTURES_DIR)
import generate_workbook  # noqa: E402


def _load_engine_module():
    """Imports src/python/engine.py as a standalone module, exactly as it
    ships (this is the same file Vite bundles via a `?raw` import for
    Pyodide, so testing it directly here tests real production code)."""
    engine_path = os.path.join(REPO_ROOT, "src", "python", "engine.py")
    spec = importlib.util.spec_from_file_location("elips_engine", engine_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class TestExcelPythonEngine(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.engine = _load_engine_module()
        cls.tmp_dir = tempfile.mkdtemp(prefix="elips_test_")
        cls.fixture_path = os.path.join(cls.tmp_dir, "elips_test_fixture.xlsx")
        generate_workbook.build_workbook().save(cls.fixture_path)
        cls.template_path = os.path.join(FIXTURES_DIR, "sample_template.md.j2")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)

    def test_01_engine_module_exposes_expected_api(self):
        """El mòdul extret ha d'exposar les funcions públiques que fa servir la resta de l'app."""
        for name in ("excel_to_json", "update_excel_from_json", "render_json_text", "render_md_two_pass_with_report"):
            self.assertTrue(hasattr(self.engine, name), f"'{name}' no exposat pel motor Python")

    def test_02_kv_sheets_header_less_and_headered_with_empty_field_cleanup(self):
        """Fulls KV sense capçalera (General) i amb capçalera Clau|Valor (pres); els camps buits queden com a ''."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]

        general = data["General"]
        self.assertEqual(general["titol_contracte"], "Subministrament d'equips informàtics")
        self.assertEqual(general["modalitat"], "Contracte Públic")
        self.assertEqual(general["nom_responsable"], "Anna Puig Soler")
        self.assertEqual(general["num_expedient"], "")
        self.assertEqual(general["data_redaccio"], "")

        pres = data["pres"]
        self.assertEqual(pres["pressupost"], "Pressupost anual")
        self.assertEqual(pres["anualitat"], 2026)

    def test_03_four_level_nested_hierarchy_incl_explicit_foreign_key(self):
        """pres -> pres.parts -> pres.parts.activitats -> pres.parts.activitats.rec (aquest darrer enllaçat
        per clau forana explícita declarada a _hierarchy_metadata, no per coincidència de nom de columna)."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]
        parts = data["pres"]["parts"]
        self.assertEqual(len(parts), 2)

        part1 = next(p for p in parts if p["id_partida"] == "PART-01")
        part2 = next(p for p in parts if p["id_partida"] == "PART-02")
        self.assertEqual(len(part1["activitats"]), 2)
        self.assertEqual(len(part2["activitats"]), 1)

        act1 = next(a for a in part1["activitats"] if a["id_activitat"] == "ACT-01")
        self.assertEqual(len(act1["rec"]), 1)
        self.assertEqual(act1["rec"][0]["recurs"], "Tècnic instal·lador")
        self.assertEqual(act1["rec"][0]["unitats"], 2)

    def test_04_ghost_rows_filtered_without_cartesian_explosion(self):
        """Les files completament buides/zero (típiques de fórmules no resoltes) es descarten, i els fills
        d'una taula plana compartida (Subcriteris) es reparteixen pel pare correcte sense duplicar-se."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]

        criteris = data["Criteris"]
        self.assertEqual(len(criteris), 3, "Les 2 files fantasma d'OUT_Criteris s'haurien d'haver descartat")

        f1 = next(c for c in criteris if c["id"] == "F1")
        v1 = next(c for c in criteris if c["id"] == "V1")
        f2 = next(c for c in criteris if c["id"] == "F2")
        self.assertEqual(len(f1["Subcriteris"]), 2)
        self.assertEqual(len(v1["Subcriteris"]), 0)
        self.assertEqual(len(f2["Subcriteris"]), 1)
        self.assertEqual(f2["Subcriteris"][0]["subid"], "F2a")

        self.assertEqual(data["Lots"], [], "OUT_Lots només conté files fantasma: ha de quedar buit")
        self.assertEqual(len(data["Mesa"]), 4)

    def test_05_update_excel_preserves_complex_formulas_as_orphans(self):
        """Les fórmules complexes (no-enllaç) mai es sobreescriuen: el nou valor es desvia al full 'orfes'."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]
        json_str = json.dumps(data, ensure_ascii=False)
        out_path = os.path.join(self.tmp_dir, "roundtrip_orphans.xlsx")

        orphan_count = self.engine.update_excel_from_json(self.fixture_path, json_str, out_path)
        self.assertEqual(orphan_count, 2, "Les 2 fórmules complexes de cost_amb_iva han de generar un orfe cada una")

        wb_out = load_workbook(out_path, data_only=False)
        self.assertIn("orfes", wb_out.sheetnames)
        ws_orfes = wb_out["orfes"]
        self.assertEqual(ws_orfes.cell(1, 1).value, "Full")
        self.assertEqual(ws_orfes.cell(1, 3).value, "Fórmula Original")
        for r in range(2, ws_orfes.max_row + 1):
            formula_val = str(ws_orfes.cell(r, 3).value or "")
            self.assertTrue(formula_val.startswith("'="), f"La fórmula orfe a la fila {r} no comença per '=: {formula_val}")

        ws_parts = wb_out["OUT_pres.parts"]
        self.assertEqual(ws_parts.cell(2, 4).value, "=C2*1.21", "La fórmula complexa no s'ha de sobreescriure")
        self.assertEqual(ws_parts.cell(2, 5).value, "=OUT_pres!B2", "L'enllaç simple s'ha de conservar intacte")

    def test_06_update_excel_leaves_untouched_sheets_byte_identical(self):
        """Un full tabular sense cap fila real (OUT_Lots) no es regenera: es deixa exactament com estava."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]
        json_str = json.dumps(data, ensure_ascii=False)
        out_path = os.path.join(self.tmp_dir, "roundtrip_untouched.xlsx")
        self.engine.update_excel_from_json(self.fixture_path, json_str, out_path)

        wb_orig = load_workbook(self.fixture_path, data_only=False)
        wb_out = load_workbook(out_path, data_only=False)
        ws_orig, ws_out = wb_orig["OUT_Lots"], wb_out["OUT_Lots"]
        self.assertEqual(ws_orig.max_row, ws_out.max_row)
        self.assertEqual(ws_orig.max_column, ws_out.max_column)
        for r in range(1, ws_orig.max_row + 1):
            for c in range(1, ws_orig.max_column + 1):
                self.assertEqual(ws_orig.cell(r, c).value, ws_out.cell(r, c).value)

    def test_07_update_excel_with_merged_cells_does_not_crash(self):
        """L'exportació sobre un full amb cel·les fusionades (OUT_Mesa, C3:C4) no ha de llançar AttributeError."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]
        json_str = json.dumps(data, ensure_ascii=False)
        out_path = os.path.join(self.tmp_dir, "roundtrip_merged.xlsx")

        orphan_count = self.engine.update_excel_from_json(self.fixture_path, json_str, out_path)
        self.assertIsInstance(orphan_count, int)
        self.assertTrue(os.path.exists(out_path))
        self.assertGreater(os.path.getsize(out_path), 0)

    def test_08_two_pass_jinja2_render_with_custom_filters_and_recovery(self):
        """Renderitza la plantilla de mostra (bucles aniuats, filtres coin/percent/words/number/prefix) i
        comprova que una variable no definida no trenca la generació (recuperació amb DebugUndefined)."""
        result = json.loads(self.engine.render_md_two_pass_with_report(self.fixture_path, self.template_path))
        self.assertTrue(result["success"], result.get("traceback"))

        md = result["markdown"]
        self.assertIn("Subministrament d'equips informàtics", md)
        self.assertIn("125.000,50", md)  # filtre 'coin'
        self.assertIn("21%", md)  # filtre 'percent'
        self.assertIn("12,00 hores", md)  # filtre 'number'
        self.assertIn("quaranta-cinc mil euros", md)  # filtre 'words'
        self.assertIn("com a President", md)  # filtre 'prefix'
        self.assertIn("Instal·lació d'equips", md)
        self.assertIn("Desplegament de programari", md)

        # La variable inexistent del final de la plantilla no ha de trencar el renderitzat
        self.assertGreater(len(result["issues"]), 0, "El motor hauria de reportar la variable indefinida com a incidència")

    def test_09_calculated_field_inside_tabular_loop_renders(self):
        """Un camp calculat (virtual, no és una columna real de l'Excel) present a l'estat viu
        (/work/in.json, equivalent a store.excelJsonData amb evaluateComputedFields ja aplicat) s'ha
        de veure reflectit en renderitzar una plantilla que itera sobre la taula que el conté."""
        data = self.engine.excel_to_json(self.fixture_path)["data"]
        for part in data["pres"]["parts"]:
            part["iva_calculat"] = round(part["import"] * 0.21, 2)
        live_json_str = json.dumps(data, ensure_ascii=False)

        loop_template_path = os.path.join(self.tmp_dir, "loop_template.md.j2")
        with open(loop_template_path, "w", encoding="utf-8") as f:
            f.write("{% for part in pres.parts %}\n- {{ part.nom_partida }}: IVA calculat = {{ part.iva_calculat }}\n{% endfor %}\n")

        live_json_path = os.path.join(self.tmp_dir, "in.json")
        with open(live_json_path, "w", encoding="utf-8") as f:
            f.write(live_json_str)

        real_exists = os.path.exists
        real_open = open

        def fake_exists(path):
            return True if path == '/work/in.json' else real_exists(path)

        def fake_open(path, *args, **kwargs):
            return real_open(live_json_path, *args, **kwargs) if path == '/work/in.json' else real_open(path, *args, **kwargs)

        with mock.patch.object(self.engine.os.path, 'exists', side_effect=fake_exists), \
             mock.patch.object(self.engine, 'open', side_effect=fake_open, create=True):
            result = json.loads(self.engine.render_md_two_pass_with_report(self.fixture_path, loop_template_path))

        self.assertTrue(result["success"], result.get("traceback"))
        md = result["markdown"]
        self.assertIn("Equips de sobretaula: IVA calculat = 9450.0", md)
        self.assertIn("Portàtils: IVA calculat = 6300.0", md)
        self.assertNotIn("iva_calculat", md, "El placeholder de recuperació indica que el camp calculat no s'ha trobat")

    def test_10_evaluate_custom_formula_mini_language(self):
        """Verifica el llenguatge de fórmules CUSTOM (SI/ARRODONEIX/CERT/FALS) que abans
        s'executava al navegador amb new Function() i ara corre dins de Pyodide."""
        ecf = self.engine.evaluate_custom_formula

        row = {"import": 100, "unitats": 3, "persones": 2, "actiu": True}

        # "import" (Catalan per "import/quantia") és un nom de camp legítim i molt
        # habitual en aquest domini: no s'ha de confondre mai amb la paraula clau
        # Python "import" (regressió real detectada i corregida durant aquesta fase).
        self.assertEqual(ecf("ARRODONEIX(import * 0.21; 2)", row), 21.0)
        self.assertEqual(ecf("import * unitats", row), 300)

        self.assertEqual(
            ecf("SI(persones > 0; persones * unitats * import; unitats * import)", row),
            600,
        )
        self.assertEqual(ecf('SI(actiu; "Sí"; "No")', row), "Sí")

        # Un enter (30) no s'ha de convertir en float (30.0) en travessar Python/JSON:
        # coherent amb com JS serialitza sempre com a "30", no "30.0".
        r = ecf("import * 3", {"import": 10})
        self.assertEqual(r, 30)
        self.assertIsInstance(r, int)

        # Una fórmula trencada no ha de petar mai: cau al valor ja present al camp.
        r = ecf("aquest_camp_no_existeix + 1", {"aquest_camp_no_existeix_backup": 5})
        self.assertEqual(r, 0)

    def test_11_evaluate_custom_formula_blocks_sandbox_escape(self):
        """La substitució de new Function() per un eval() Python restringit s'ha de mantenir
        tancada als intents habituals d'escapar del sandbox (accés a dunders / __import__)."""
        ecf = self.engine.evaluate_custom_formula
        self.assertEqual(ecf("().__class__.__bases__[0].__subclasses__()", {}), 0)
        self.assertEqual(ecf("__import__('os').system('echo pwned')", {}), 0)

    def test_12_evaluate_computed_fields_custom_and_aggregation(self):
        """Verifica el motor complet (evaluate_computed_fields): una fórmula CUSTOM per
        fila i una agregació SUM sobre la mateixa taula, tal com fa servir l'aplicació."""
        data = {
            "pres": {
                "anualitat": 2026,
                "parts": [
                    {"id": "A", "import": 100, "iva": None},
                    {"id": "B", "import": 200, "iva": None},
                ],
                "total": None,
            }
        }
        metadata = [
            {"group": "pres.parts", "element": "iva", "type": "Computed", "calcFormula": "ARRODONEIX(import * 0.21; 2)"},
            {"group": "pres", "element": "total", "type": "Computed", "calcFn": "SUM", "calcVector": "parts", "calcTargetCol": "import"},
        ]
        result = json.loads(self.engine.evaluate_computed_fields(json.dumps(data), json.dumps(metadata)))
        self.assertTrue(result["success"])
        parts = result["data"]["pres"]["parts"]
        self.assertEqual(parts[0]["iva"], 21.0)
        self.assertEqual(parts[1]["iva"], 42.0)
        self.assertEqual(result["data"]["pres"]["total"], 300)

    def test_13_trim_blocks_avoids_blank_lines_between_loop_rows(self):
        """trim_blocks/lstrip_blocks=True (the Jinja2 Environment config in engine.py) must
        strip a block tag's own trailing newline -- otherwise every {% for %}/{% endfor %}
        (and {% if %}/{% endif %}) leaves a blank line behind in the rendered output, which
        breaks a Markdown table generated by looping over its rows (a blank line ends a
        table) and adds unwanted spacing around conditionally-included paragraphs."""
        table_template_path = os.path.join(self.tmp_dir, "table_template.md.j2")
        with open(table_template_path, "w", encoding="utf-8") as f:
            f.write(
                "| Partida | Import |\n"
                "| --- | ---: |\n"
                "{% for part in pres.parts %}\n"
                "| {{ part.nom_partida }} | {{ part.import }} |\n"
                "{% endfor %}\n"
                "{% if pres.parts %}\n"
                "Hi ha partides.\n"
                "{% endif %}\n"
            )
        result = json.loads(self.engine.render_md_two_pass_with_report(self.fixture_path, table_template_path))
        self.assertTrue(result["success"], result.get("traceback"))
        md = result["markdown"]

        self.assertNotIn('\n\n', md.strip(), "Una línia en blanc enmig trencaria la taula Markdown: " + repr(md))
        lines = [l for l in md.splitlines() if l.strip()]
        table_lines, rest_lines = lines[:-1], lines[-1:]
        self.assertTrue(all(l.startswith('|') for l in table_lines), md)
        self.assertEqual(rest_lines, ['Hi ha partides.'])


if __name__ == "__main__":
    unittest.main()
